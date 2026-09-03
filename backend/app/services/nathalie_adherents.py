"""
Annuaire adhérents Union pour Nathalie (création / consultation).

Table Supabase `nathalie_adherents` : source de vérité à la place de LISTE CLIENT 2 (Sheets).
Drive (pièces) et Gmail (ouvertures fournisseurs) restent gérés dans nathalie_service.
"""
from __future__ import annotations

import re
import unicodedata
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy import text

from app.database import engine

TABLE = "nathalie_adherents"

COLUMNS = [
    "id",
    "code_union",
    "nom_client",
    "groupe",
    "contact_magasin",
    "telephone",
    "mail",
    "adresse",
    "code_postal",
    "ville",
    "departement",
    "siret",
    "tva",
    "raison_sociale",
    "etat_insee",
    "perimetre",
    "region_commerciale",
    "notes",
    "is_closed",
    "agent_union",
    "contrat_union",
    "ouverture_chez",
    "contact_responsable_pdv",
    "telephone_responsable",
    "contact_appro",
    "rib_url",
    "kbis_url",
    "piece_identite_url",
    "photo_devanture_url",
    "photo_comptoir_url",
    "photo_stock_url",
    "photo_autre_1_url",
    "photo_autre_2_url",
    "drive_folder_id",
    "drive_link",
    "drive_checked_at",
    "date_creation_compte",
    "source",
    "created_at",
    "updated_at",
]


def _is_pg() -> bool:
    return engine.dialect.name == "postgresql"


def _now():
    return datetime.utcnow() if _is_pg() else datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _uuid() -> str:
    return str(uuid.uuid4())


def _bool(val) -> Any:
    if _is_pg():
        return bool(val)
    return 1 if val else 0


def _coerce_closed(val, nom: str, notes: Optional[str], etat_insee) -> Any:
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return _bool(is_closed_from(nom, notes, etat_insee))
    if isinstance(val, str):
        return _bool(val.strip().lower() in ("1", "true", "oui", "on", "yes"))
    return _bool(val)


def ensure_tables() -> None:
    if _is_pg():
        ddl = f'''
        CREATE TABLE IF NOT EXISTS {TABLE} (
          id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
          code_union text NOT NULL UNIQUE,
          nom_client text NOT NULL DEFAULT '',
          groupe text,
          contact_magasin text,
          telephone text,
          mail text,
          adresse text,
          code_postal text,
          ville text,
          departement text,
          siret text,
          tva text,
          raison_sociale text,
          etat_insee text,
          perimetre text,
          region_commerciale text,
          notes text,
          is_closed boolean NOT NULL DEFAULT false,
          agent_union text,
          contrat_union text,
          ouverture_chez text,
          contact_responsable_pdv text,
          telephone_responsable text,
          contact_appro text,
          rib_url text,
          kbis_url text,
          piece_identite_url text,
          photo_devanture_url text,
          photo_comptoir_url text,
          photo_stock_url text,
          photo_autre_1_url text,
          photo_autre_2_url text,
          drive_folder_id text,
          drive_link text,
          drive_checked_at timestamptz,
          date_creation_compte timestamptz,
          source text NOT NULL DEFAULT 'manuel',
          created_at timestamptz NOT NULL DEFAULT now(),
          updated_at timestamptz NOT NULL DEFAULT now()
        )
        '''
        comments = [
            f"COMMENT ON TABLE {TABLE} IS "
            "'Annuaire adhérents Union pour Nathalie (création de comptes). Remplace LISTE CLIENT 2 Sheets.'",
        ]
    else:
        ddl = f'''
        CREATE TABLE IF NOT EXISTS {TABLE} (
          id TEXT PRIMARY KEY,
          code_union TEXT NOT NULL UNIQUE,
          nom_client TEXT NOT NULL DEFAULT '',
          groupe TEXT,
          contact_magasin TEXT,
          telephone TEXT,
          mail TEXT,
          adresse TEXT,
          code_postal TEXT,
          ville TEXT,
          departement TEXT,
          siret TEXT,
          tva TEXT,
          raison_sociale TEXT,
          etat_insee TEXT,
          perimetre TEXT,
          region_commerciale TEXT,
          notes TEXT,
          is_closed INTEGER NOT NULL DEFAULT 0,
          agent_union TEXT,
          contrat_union TEXT,
          ouverture_chez TEXT,
          contact_responsable_pdv TEXT,
          telephone_responsable TEXT,
          contact_appro TEXT,
          rib_url TEXT,
          kbis_url TEXT,
          piece_identite_url TEXT,
          photo_devanture_url TEXT,
          photo_comptoir_url TEXT,
          photo_stock_url TEXT,
          photo_autre_1_url TEXT,
          photo_autre_2_url TEXT,
          drive_folder_id TEXT,
          drive_link TEXT,
          drive_checked_at TEXT,
          date_creation_compte TEXT,
          source TEXT NOT NULL DEFAULT 'manuel',
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        '''
        comments = []
    indexes = [
        f"CREATE INDEX IF NOT EXISTS idx_nathalie_adherents_groupe ON {TABLE}(groupe)",
        f"CREATE INDEX IF NOT EXISTS idx_nathalie_adherents_ville ON {TABLE}(ville)",
        f"CREATE INDEX IF NOT EXISTS idx_nathalie_adherents_closed ON {TABLE}(is_closed)",
        f"CREATE INDEX IF NOT EXISTS idx_nathalie_adherents_region ON {TABLE}(region_commerciale)",
        f"CREATE INDEX IF NOT EXISTS idx_nathalie_adherents_agent ON {TABLE}(agent_union)",
    ]
    extra_pg = [
        "departement text",
        "raison_sociale text",
        "etat_insee text",
        "perimetre text",
        "region_commerciale text",
        "contact_responsable_pdv text",
        "telephone_responsable text",
        "contact_appro text",
        "photo_devanture_url text",
        "photo_comptoir_url text",
        "photo_stock_url text",
        "photo_autre_1_url text",
        "photo_autre_2_url text",
        "drive_checked_at timestamptz",
        "date_creation_compte timestamptz",
    ]
    extra_sqlite = [
        "departement",
        "raison_sociale",
        "etat_insee",
        "perimetre",
        "region_commerciale",
        "contact_responsable_pdv",
        "telephone_responsable",
        "contact_appro",
        "photo_devanture_url",
        "photo_comptoir_url",
        "photo_stock_url",
        "photo_autre_1_url",
        "photo_autre_2_url",
        "drive_checked_at",
        "date_creation_compte",
    ]
    with engine.begin() as conn:
        conn.execute(text(ddl))
        for stmt in comments:
            try:
                conn.execute(text(stmt))
            except Exception:
                pass
        try:
            if _is_pg():
                for col in extra_pg:
                    conn.execute(text(f"ALTER TABLE {TABLE} ADD COLUMN IF NOT EXISTS {col}"))
            else:
                cols = conn.execute(text(f"PRAGMA table_info({TABLE})")).fetchall()
                names = {c[1] for c in cols}
                for col in extra_sqlite:
                    if col not in names:
                        conn.execute(text(f"ALTER TABLE {TABLE} ADD COLUMN {col} TEXT"))
        except Exception:
            pass
        for stmt in indexes:
            try:
                conn.execute(text(stmt))
            except Exception:
                pass
        if _is_pg():
            try:
                conn.execute(text(f"ALTER TABLE {TABLE} ENABLE ROW LEVEL SECURITY"))
            except Exception:
                pass


def clean_code_union(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().upper()
    if s in ("", "?", "-", "N/A", "NA"):
        return None
    return s


def clean_postal(value: Any) -> Optional[str]:
    if value is None or str(value).strip() in ("", "-", "?"):
        return None
    if isinstance(value, float):
        return f"{int(value):05d}"
    s = str(value).strip()
    digits = re.search(r"(\d{5})", s.replace(" ", ""))
    if digits:
        return digits.group(1)
    compact = s.replace(" ", "")
    try:
        return f"{int(float(compact)):05d}"
    except (ValueError, OverflowError):
        return s or None


def clean_siret(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float):
        return str(int(value))
    s = str(value).strip().replace(" ", "")
    if s in ("", "?", "-", "N/A"):
        return None
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "?", "-", "N/A", "NA", "None"):
        return None
    return s


def is_closed_from(nom: str, notes: Optional[str] = None, etat_insee: Optional[str] = None) -> bool:
    blob = f"{nom or ''} {notes or ''} {etat_insee or ''}".upper()
    return any(k in blob for k in ("FERME", "FERMÉ", "LIQUIDATION", "RADIE", "CESSÉE", "CESSEE"))


def _row_to_client(row) -> Dict[str, Any]:
    data = dict(row)
    for key in ("created_at", "updated_at", "drive_checked_at", "date_creation_compte"):
        val = data.get(key)
        if hasattr(val, "isoformat"):
            data[key] = val.isoformat()
    closed = data.get("is_closed")
    data["is_closed"] = bool(closed)
    rib = data.get("rib_url") or ""
    kbis = data.get("kbis_url") or ""
    piece = data.get("piece_identite_url") or ""
    data["rib"] = rib
    data["kbis"] = kbis
    data["piece_identite"] = piece
    data["gerant"] = data.get("contact_magasin") or ""
    data["gerant_telephone"] = data.get("telephone") or ""
    data["gerant_mail"] = data.get("mail") or ""
    data["responsable_magasin"] = data.get("contact_responsable_pdv") or ""
    data["photos"] = {
        "devanture": data.get("photo_devanture_url") or "",
        "comptoir": data.get("photo_comptoir_url") or "",
        "stock": data.get("photo_stock_url") or "",
        "autre_1": data.get("photo_autre_1_url") or "",
        "autre_2": data.get("photo_autre_2_url") or "",
    }
    data["note_generale"] = data.get("notes") or ""
    data["docs_complets"] = bool(rib and kbis and piece)
    data["has_rib"] = bool(rib)
    data["has_kbis"] = bool(kbis)
    data["drive_checked"] = bool(data.get("drive_checked_at") or data.get("drive_folder_id"))
    missing = []
    if not rib:
        missing.append("RIB")
    if not kbis:
        missing.append("Kbis")
    data["missing_docs"] = missing
    data["dossier_complet"] = not missing
    data["dossier_en_cours"] = bool(data["drive_checked"] and missing)
    return data


def list_clients(with_ouverture_only: bool = False) -> List[Dict[str, Any]]:
    ensure_tables()
    clauses = []
    if with_ouverture_only:
        clauses.append("ouverture_chez IS NOT NULL AND TRIM(ouverture_chez) <> ''")
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"SELECT * FROM {TABLE}{where} ORDER BY nom_client ASC"
    with engine.connect() as conn:
        rows = conn.execute(text(sql)).mappings().all()
    return [_row_to_client(r) for r in rows]


def get_by_code(code_union: str) -> Optional[Dict[str, Any]]:
    ensure_tables()
    code = clean_code_union(code_union)
    if not code:
        return None
    with engine.connect() as conn:
        row = conn.execute(
            text(f"SELECT * FROM {TABLE} WHERE UPPER(code_union) = :code"),
            {"code": code},
        ).mappings().first()
    return _row_to_client(row) if row else None


def next_code_union(prefix: str = "M") -> str:
    """Prochain code Mxxxx d'après l'annuaire (liste clients), pas Drive."""
    ensure_tables()
    prefix = (prefix or "M").upper()
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT code_union FROM {TABLE} WHERE UPPER(code_union) LIKE :p"),
            {"p": f"{prefix}%"},
        ).fetchall()
    max_n = 0
    for (code,) in rows:
        s = str(code or "").strip().upper()
        if s.startswith(prefix) and s[len(prefix):].isdigit():
            max_n = max(max_n, int(s[len(prefix):]))
    return f"{prefix}{max_n + 1:04d}"


def normalize_groupe_label(groupe: Optional[str]) -> Optional[str]:
    s = (groupe or "").strip()
    if not s:
        return None
    u = s.upper()
    if u in ("INDEPENDANT", "INDÉPENDANT", "INDEPENDANT UNION"):
        return "INDEPENDANT UNION"
    return s


def upsert_client(payload: Dict[str, Any], *, keep_docs: bool = True) -> Dict[str, Any]:
    ensure_tables()
    code = clean_code_union(payload.get("code_union"))
    if not code:
        raise ValueError("code_union requis")
    existing = get_by_code(code)
    nom = (payload.get("nom_client") or "").strip()
    notes = clean_text(payload.get("notes") if payload.get("notes") is not None else payload.get("note_generale"))
    record = {
        "code_union": code,
        "nom_client": nom,
        "groupe": normalize_groupe_label(payload.get("groupe")),
        "contact_magasin": clean_text(payload.get("contact_magasin") or payload.get("contact")),
        "telephone": clean_text(payload.get("telephone")),
        "mail": clean_text(payload.get("mail")),
        "adresse": clean_text(payload.get("adresse")),
        "code_postal": clean_postal(payload.get("code_postal")),
        "ville": clean_text(payload.get("ville")),
        "departement": clean_text(payload.get("departement")),
        "siret": clean_siret(payload.get("siret")),
        "tva": clean_text(payload.get("tva")),
        "raison_sociale": clean_text(payload.get("raison_sociale")),
        "etat_insee": clean_text(payload.get("etat_insee")),
        "perimetre": clean_text(payload.get("perimetre")),
        "region_commerciale": clean_text(
            payload.get("region_commerciale") or payload.get("region")
        ),
        "notes": notes,
        "is_closed": _coerce_closed(payload.get("is_closed"), nom, notes, payload.get("etat_insee")),
        "agent_union": _canonical_agent(payload.get("agent_union")),
        "contrat_union": clean_text(payload.get("contrat_union") or payload.get("contrat_type")),
        "ouverture_chez": clean_text(payload.get("ouverture_chez")),
        "contact_responsable_pdv": clean_text(
            payload.get("contact_responsable_pdv") or payload.get("responsable_magasin")
        ),
        "telephone_responsable": clean_text(payload.get("telephone_responsable")),
        "contact_appro": clean_text(payload.get("contact_appro") or payload.get("contact_achat")),
        "rib_url": clean_text(payload.get("rib_url") or payload.get("rib")),
        "kbis_url": clean_text(payload.get("kbis_url") or payload.get("kbis")),
        "piece_identite_url": clean_text(payload.get("piece_identite_url") or payload.get("piece_identite")),
        "photo_devanture_url": clean_text(payload.get("photo_devanture_url") or payload.get("photo_devanture")),
        "photo_comptoir_url": clean_text(payload.get("photo_comptoir_url") or payload.get("photo_comptoir")),
        "photo_stock_url": clean_text(payload.get("photo_stock_url") or payload.get("photo_stock")),
        "photo_autre_1_url": clean_text(payload.get("photo_autre_1_url") or payload.get("photo_autre_1")),
        "photo_autre_2_url": clean_text(payload.get("photo_autre_2_url") or payload.get("photo_autre_2")),
        "drive_folder_id": clean_text(payload.get("drive_folder_id")),
        "drive_link": clean_text(payload.get("drive_link")),
        "drive_checked_at": payload.get("drive_checked_at"),
        "date_creation_compte": payload.get("date_creation_compte"),
        "source": payload.get("source") or "manuel",
        "updated_at": _now(),
    }
    if existing and keep_docs:
        aliases = {
            "rib_url": existing.get("rib_url") or existing.get("rib"),
            "kbis_url": existing.get("kbis_url") or existing.get("kbis"),
            "piece_identite_url": existing.get("piece_identite_url") or existing.get("piece_identite"),
            "photo_devanture_url": existing.get("photo_devanture_url"),
            "photo_comptoir_url": existing.get("photo_comptoir_url"),
            "photo_stock_url": existing.get("photo_stock_url"),
            "photo_autre_1_url": existing.get("photo_autre_1_url"),
            "photo_autre_2_url": existing.get("photo_autre_2_url"),
            "drive_folder_id": existing.get("drive_folder_id"),
            "drive_link": existing.get("drive_link"),
            "drive_checked_at": existing.get("drive_checked_at"),
            "date_creation_compte": existing.get("date_creation_compte"),
            "ouverture_chez": existing.get("ouverture_chez"),
            "contrat_union": existing.get("contrat_union"),
        }
        for key, previous in aliases.items():
            if not record.get(key) and previous:
                record[key] = previous

    now = _now()
    if existing:
        sets = ", ".join(f"{c} = :{c}" for c in COLUMNS if c not in ("id", "code_union", "created_at", "date_creation_compte"))
        params = {k: record[k] for k in COLUMNS if k not in ("id", "code_union", "created_at", "date_creation_compte")}
        params["code"] = code
        with engine.begin() as conn:
            conn.execute(
                text(f"UPDATE {TABLE} SET {sets} WHERE UPPER(code_union) = :code"),
                params,
            )
        updated = get_by_code(code)
        if not updated:
            raise RuntimeError("Mise à jour adhérent introuvable")
        return updated

    record["id"] = payload.get("id") or _uuid()
    record["created_at"] = now
    if not record.get("date_creation_compte") and (record.get("source") or "manuel") == "manuel":
        record["date_creation_compte"] = now
    cols = ", ".join(COLUMNS)
    placeholders = ", ".join(f":{c}" for c in COLUMNS)
    with engine.begin() as conn:
        conn.execute(text(f"INSERT INTO {TABLE} ({cols}) VALUES ({placeholders})"), record)
    created = get_by_code(code)
    if not created:
        raise RuntimeError("Adhérent créé mais introuvable")
    return created


def _norm_header(value: Any) -> str:
    s = unicodedata.normalize("NFKD", str(value or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


_EXCEL_HEADERS = {
    "code union": "code_union",
    "perimetre": "perimetre",
    "nom client": "nom_client",
    "magasins": "nom_client",
    "groupe": "groupe",
    "region commercial": "region_commerciale",
    "contact magasin": "contact_magasin",
    "adresse": "adresse",
    "code postal": "code_postal",
    "departement": "departement",
    "ville": "ville",
    "telephone": "telephone",
    "contact responsable pdv": "contact_responsable_pdv",
    "contact achat appro": "contact_appro",
    "mail": "mail",
    "agent union": "agent_union",
    "siret": "siret",
    "tva": "tva",
    "raison sociale officielle": "raison_sociale",
    "etat insee": "etat_insee",
    "notes": "notes",
    "commentaires": "notes",
}


def _pick_excel_sheet(wb):
    for name in wb.sheetnames:
        if "liste client" in name.lower():
            return wb[name]
    for name in wb.sheetnames:
        if "juillet" in name.lower():
            return wb[name]
    return wb[wb.sheetnames[-1]]


def _header_map(row) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, cell in enumerate(row):
        key = _EXCEL_HEADERS.get(_norm_header(cell))
        if key:
            mapping[idx] = key
    return mapping


def parse_excel_rows(path: str) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = _pick_excel_sheet(wb)

    header_row_idx = None
    col_map: Dict[int, str] = {}
    for i, raw in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        mapping = _header_map(raw)
        if "code_union" in mapping.values() and "nom_client" in mapping.values():
            header_row_idx = i
            col_map = mapping
            break

    if header_row_idx is None:
        return _parse_juillet_fixed(sheet)

    rows: List[Dict[str, Any]] = []
    by_code: Dict[str, Dict[str, Any]] = {}
    for raw in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        parsed: Dict[str, Any] = {}
        for idx, key in col_map.items():
            parsed[key] = raw[idx] if idx < len(raw) else None
        nom = clean_text(parsed.get("nom_client")) or ""
        code = clean_code_union(parsed.get("code_union"))
        if not nom and not code:
            continue
        notes = clean_text(parsed.get("notes"))
        etat = clean_text(parsed.get("etat_insee"))
        row = {
            "code_union": code,
            "nom_client": nom,
            "groupe": parsed.get("groupe"),
            "contact_magasin": parsed.get("contact_magasin"),
            "telephone": parsed.get("telephone"),
            "mail": parsed.get("mail"),
            "adresse": parsed.get("adresse"),
            "code_postal": parsed.get("code_postal"),
            "ville": parsed.get("ville"),
            "departement": parsed.get("departement"),
            "siret": parsed.get("siret"),
            "tva": parsed.get("tva"),
            "raison_sociale": parsed.get("raison_sociale"),
            "etat_insee": etat,
            "perimetre": parsed.get("perimetre"),
            "region_commerciale": parsed.get("region_commerciale"),
            "contact_responsable_pdv": parsed.get("contact_responsable_pdv"),
            "contact_appro": parsed.get("contact_appro"),
            "agent_union": parsed.get("agent_union"),
            "notes": notes,
            "is_closed": is_closed_from(nom, notes, etat),
            "source": "excel",
        }
        if code:
            by_code[code] = row
        else:
            rows.append(row)
    return rows + list(by_code.values())


def _parse_juillet_fixed(sheet) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=2, max_col=12, values_only=True):
        vals = list(raw) + [None] * 12
        code, mag, groupe, contact, tel, mail, adr, cp, ville, siret, tva, notes = vals[:12]
        if not mag and not code:
            continue
        nom = str(mag).strip() if mag else ""
        notes_s = clean_text(notes)
        rows.append({
            "code_union": code,
            "nom_client": nom,
            "groupe": groupe,
            "contact_magasin": contact,
            "telephone": tel,
            "mail": mail,
            "adresse": adr,
            "code_postal": cp,
            "ville": ville,
            "siret": siret,
            "tva": tva,
            "notes": notes_s,
            "is_closed": is_closed_from(nom, notes_s),
            "source": "excel",
        })
    return rows


def import_excel_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    created = 0
    updated = 0
    errors: List[str] = []
    for i, raw in enumerate(rows, start=1):
        try:
            code = clean_code_union(raw.get("code_union"))
            if not code:
                continue
            existed = get_by_code(code) is not None
            upsert_client(raw, keep_docs=True)
            if existed:
                updated += 1
            else:
                created += 1
        except Exception as exc:
            errors.append(f"Ligne {i}: {exc}")
    return {"created": created, "updated": updated, "errors": errors[:30], "total": created + updated}


def replace_directory(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Remplace l'annuaire par l'Excel : upsert + suppression des codes absents.

    Conserve RIB / Kbis / pièce / dossier Drive déjà liés au même code_union.
    """
    stats = import_excel_rows(rows)
    imported = {clean_code_union(r.get("code_union")) for r in rows}
    imported.discard(None)
    deleted = 0
    if imported:
        ensure_tables()
        with engine.begin() as conn:
            existing = [
                clean_code_union(r[0])
                for r in conn.execute(text(f"SELECT code_union FROM {TABLE}")).fetchall()
            ]
            to_del = [c for c in existing if c and c not in imported]
            for i in range(0, len(to_del), 80):
                chunk = to_del[i : i + 80]
                params = {f"c{j}": c for j, c in enumerate(chunk)}
                placeholders = ", ".join(f":c{j}" for j in range(len(chunk)))
                conn.execute(
                    text(f"DELETE FROM {TABLE} WHERE UPPER(code_union) IN ({placeholders})"),
                    params,
                )
            deleted = len(to_del)
    stats["deleted"] = deleted
    return stats


def _canonical_agent(value: Any) -> Optional[str]:
    from app.services.commercial_scope import canonical_commercial_label
    return canonical_commercial_label(clean_text(value))


def reassign_code_union_portfolio(
    code_union: str,
    commercial: Optional[str],
    region: Optional[str] = None,
) -> Dict[str, int]:
    """Réaffecte le CA / analyses Pure Data et les impayés au nouvel agent Union."""
    code = clean_code_union(code_union)
    if not code:
        return {}
    comm = _canonical_agent(commercial)
    region_s = clean_text(region)
    counts: Dict[str, int] = {}
    targets = (
        ("pure_data_monthly", True),
        ("pure_data_cumulative", True),
        ("pure_data", True),
        ("impayes", False),
    )
    for table, has_region in targets:
        sets = ["commercial = :c"] if comm is not None else []
        params: Dict[str, Any] = {"code": code}
        if comm is not None:
            params["c"] = comm
        if has_region and region_s:
            sets.append("region_commerciale = :r")
            params["r"] = region_s
        if not sets:
            continue
        sql = (
            f'UPDATE "{table}" SET {", ".join(sets)} '
            "WHERE UPPER(TRIM(code_union)) = :code"
        )
        try:
            with engine.begin() as conn:
                result = conn.execute(text(sql), params)
            counts[table] = int(result.rowcount or 0)
        except Exception:
            counts[table] = 0
    return counts


def patch_drive_docs(code_union: str, fields: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Met à jour uniquement les liens Drive / pièces, sans toucher au reste de la fiche."""
    ensure_tables()
    existing = get_by_code(code_union)
    if not existing:
        return None
    allowed = {
        "rib_url",
        "kbis_url",
        "piece_identite_url",
        "photo_devanture_url",
        "photo_comptoir_url",
        "photo_stock_url",
        "photo_autre_1_url",
        "photo_autre_2_url",
        "drive_folder_id",
        "drive_link",
        "drive_checked_at",
    }
    updates = {k: v for k, v in fields.items() if k in allowed and v not in (None, "")}
    if not updates:
        return existing
    updates["updated_at"] = _now()
    sets = ", ".join(f"{k} = :{k}" for k in updates)
    updates["code"] = clean_code_union(code_union)
    with engine.begin() as conn:
        conn.execute(
            text(f"UPDATE {TABLE} SET {sets} WHERE UPPER(code_union) = :code"),
            updates,
        )
    return get_by_code(code_union)
