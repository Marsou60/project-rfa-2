"""
Routes API FastAPI.
"""
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, UploadFile, File, HTTPException, Query, Depends, Header, Form, Body
from fastapi.responses import JSONResponse, Response, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlmodel import Session, select
from datetime import datetime
import tempfile
import os
import json
import uuid
import base64
from app.services.excel_import import load_excel
from app.services.compute import (
    get_clients_summary, 
    get_client_detail, 
    get_group_detail,
    get_entity_detail_with_rfa,
    compute_aggregations,
    get_global_recap_rfa
)
from app.services.contract_resolver import resolve_contract
from app.services.rfa_calculator import calculate_rfa
from app.services.pdf_export import generate_pdf_report
from app.storage import (
    create_import,
    get_import,
    list_imports,
    set_live_import,
    get_live_import,
    LIVE_IMPORT_ID,
    create_pure_data_import,
    get_pure_data_import,
)
from app.schemas import (
    UploadResponse,
    SyncFromSheetsRequest,
    ClientSummary,
    ClientDetail,
    GroupDetail,
    EntitySummary,
    EntityDetailWithRfa,
    RecapGlobalRfa,
    AdCreate,
    AdUpdate,
    AdResponse,
    LoginRequest,
    LoginResponse,
    UserCreate,
    UserUpdate,
    UserResponse,
)
from app.database import get_session, hash_password, verify_password, UPLOADS_DIR, AVATARS_DIR, LOGOS_DIR, SUPPLIER_LOGOS_DIR
from app.models import Contract, ContractRule, ContractAssignment, ContractOverride, RuleScope, TargetType, OverrideTierType, Ad, User, UserRole, AppSettings, SupplierLogo
from app.services import nathalie_service

router = APIRouter()


@router.post("/upload", response_model=UploadResponse)
async def upload_excel(file: UploadFile = File(...)):
    """Upload et import d'un fichier Excel."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Le fichier doit être un .xlsx ou .xls")
    
    tmp_path = None
    try:
        # Sauvegarder temporairement
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        # Importer
        data, raw_columns, column_mapping = load_excel(tmp_path)
        
        # Vérifier qu'on a des données
        if not data:
            raise HTTPException(
                status_code=400,
                detail="Aucune donnée valide trouvée dans le fichier. Vérifiez que les colonnes 'Code Union' et 'Groupe Client' sont présentes et remplies."
            )
        
        # Créer l'import
        import_id = create_import(raw_columns, column_mapping, data)
        
        # Calculer les agrégations
        import_data = get_import(import_id)
        if import_data:
            try:
                compute_aggregations(import_data)
            except Exception as agg_error:
                import traceback
                print(f"Erreur lors de l'agrégation: {agg_error}")
                print(traceback.format_exc())
                raise HTTPException(
                    status_code=500,
                    detail=f"Erreur lors de l'agrégation des données: {str(agg_error)}"
                )
        
        return UploadResponse(
            import_id=import_id,
            meta={
                "filename": file.filename,
                "nb_lignes": len(data),
            },
            nb_lignes=len(data),
            colonnes_brutes=raw_columns,
            colonnes_reconnues=column_mapping,
        )
    except Exception as e:
        import traceback
        error_detail = str(e)
        traceback_str = traceback.format_exc()
        print(f"Erreur lors de l'upload: {error_detail}")
        print(traceback_str)
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors du traitement du fichier: {error_detail}"
        )
    finally:
        # Nettoyer
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@router.post("/sync-from-sheets", response_model=UploadResponse)
async def sync_from_sheets(body: SyncFromSheetsRequest):
    """
    Synchronise les données RFA depuis un Google Sheet (même structure que l'Excel).
    Nécessite les dépendances optionnelles (pip install -r requirements-sheets.txt)
    et GOOGLE_APPLICATION_CREDENTIALS pointant vers le JSON du compte de service.
    """
    try:
        from app.services.sheets_loader import load_from_sheets
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail="Google Sheets non configuré. Installez : pip install -r requirements-sheets.txt",
        ) from e
    try:
        data, raw_columns, column_mapping = load_from_sheets(
            spreadsheet_id=body.spreadsheet_id,
            sheet_name_or_range=body.sheet_name,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur Google Sheets: {str(e)}")
    if not data:
        raise HTTPException(
            status_code=400,
            detail="Aucune donnée valide dans le Sheet. Vérifiez la première feuille (en-têtes + lignes avec Code Union).",
        )
    import_id = create_import(raw_columns, column_mapping, data)
    import_data = get_import(import_id)
    if import_data:
        try:
            compute_aggregations(import_data)
        except Exception as agg_error:
            import traceback
            print(traceback.format_exc())
            raise HTTPException(
                status_code=500,
                detail=f"Erreur lors de l'agrégation: {str(agg_error)}",
            )
    return UploadResponse(
        import_id=import_id,
        meta={"source": "google_sheets", "spreadsheet_id": body.spreadsheet_id, "nb_lignes": len(data)},
        nb_lignes=len(data),
        colonnes_brutes=raw_columns,
        colonnes_reconnues=column_mapping,
    )


# ── RFA Sheets (feuille connectée = source de données pour tous) ─────────────────

def _get_rfa_sheets_config(session: Optional[Session] = None) -> tuple:
    """Retourne (spreadsheet_id, sheet_name) depuis AppSettings puis .env."""
    sid, sname = None, None
    if session:
        st = session.exec(select(AppSettings).where(AppSettings.key == "rfa_sheets_spreadsheet_id")).first()
        if st and st.value:
            sid = st.value
        st = session.exec(select(AppSettings).where(AppSettings.key == "rfa_sheets_sheet_name")).first()
        if st and st.value:
            sname = st.value
    if not sid:
        sid = os.environ.get("RFA_SHEETS_SPREADSHEET_ID", "").strip()
    if sname is None or sname == "":
        sname = os.environ.get("RFA_SHEETS_SHEET_NAME", "").strip() or None
    return (sid or None, sname or None)


_CACHE_KEY_DATA   = "sheets_live_raw_data"
_CACHE_KEY_COLS   = "sheets_live_raw_columns"
_CACHE_KEY_MAP    = "sheets_live_col_mapping"
_CACHE_KEY_CLIENT = "sheets_live_by_client"
_CACHE_KEY_GROUP  = "sheets_live_by_group"
_CACHE_KEY_GENIE  = "sheets_live_genie_analysis"


def _upsert_setting(session: Session, key: str, value: str):
    st = session.exec(select(AppSettings).where(AppSettings.key == key)).first()
    if st:
        st.value = value
        st.updated_at = datetime.now()
    else:
        session.add(AppSettings(key=key, value=value))


def _save_live_cache(session: Session, data_list: list, raw_columns: list, column_mapping: dict):
    """Sauvegarde le cache complet (raw + agrégations) dans Supabase."""
    try:
        import json as _json
        # 1. Données brutes + colonnes
        _upsert_setting(session, _CACHE_KEY_DATA, _json.dumps(data_list))
        _upsert_setting(session, _CACHE_KEY_COLS, _json.dumps(raw_columns))
        _upsert_setting(session, _CACHE_KEY_MAP,  _json.dumps(column_mapping))
        # 2. Données agrégées (by_client, by_group) — calculées maintenant pour éviter de le refaire à chaque requête
        set_live_import(raw_columns, column_mapping, data_list)
        import_data = get_live_import()
        if import_data:
            try:
                compute_aggregations(import_data)
                _upsert_setting(session, _CACHE_KEY_CLIENT, _json.dumps(import_data.by_client))
                _upsert_setting(session, _CACHE_KEY_GROUP,  _json.dumps(import_data.by_group))
            except Exception as e:
                print(f"[CACHE] Erreur agrégations: {e}")
        session.commit()
    except Exception as e:
        print(f"[CACHE] Erreur sauvegarde: {e}")


def _load_live_cache(session: Optional[Session]):
    """Charge le cache depuis Supabase — agrégations pré-calculées incluses."""
    if not session:
        return None
    try:
        import json as _json
        def _get(key):
            st = session.exec(select(AppSettings).where(AppSettings.key == key)).first()
            return _json.loads(st.value) if st and st.value else None

        data_list      = _get(_CACHE_KEY_DATA)
        raw_columns    = _get(_CACHE_KEY_COLS)
        column_mapping = _get(_CACHE_KEY_MAP)
        by_client      = _get(_CACHE_KEY_CLIENT)
        by_group       = _get(_CACHE_KEY_GROUP)

        if not data_list or not raw_columns:
            return None
        set_live_import(raw_columns, column_mapping or {}, data_list)
        import_data = get_live_import()
        if import_data:
            if by_client:
                import_data.by_client = by_client
            if by_group:
                import_data.by_group = by_group
            # Si pas d'agrégations en cache, les calculer (lent, mais seulement 1 fois)
            if not by_client or not by_group:
                try:
                    compute_aggregations(import_data)
                except Exception:
                    pass
        return import_data
    except Exception as e:
        print(f"[CACHE] Erreur lecture: {e}")
        return None


def _resolve_import_data(import_id: str, session: Optional[Session] = None):
    """
    Résout import_id en ImportData.
    Pour sheets_live : mémoire → table rfa_data (Supabase) → rechargement depuis Sheets.
    """
    if import_id != LIVE_IMPORT_ID:
        return get_import(import_id)
    # 1. Mémoire (même process Vercel)
    data = get_live_import()
    if data:
        return data
    # 2. Table rfa_data Supabase (rapide, survit aux cold starts)
    if session:
        try:
            from app.services.rfa_supabase import read_rfa_from_supabase, build_column_mapping
            data_list = read_rfa_from_supabase(session)
            if data_list:
                column_mapping = build_column_mapping()
                raw_columns = list(column_mapping.keys())
                set_live_import(raw_columns, column_mapping, data_list)
                import_data = get_live_import()
                if import_data:
                    try:
                        compute_aggregations(import_data)
                    except Exception:
                        pass
                return import_data
        except Exception as e:
            print(f"[RESOLVE] Erreur lecture rfa_data: {e}")
    # 3. Rechargement depuis Google Sheets (lent, dernier recours)
    spreadsheet_id, sheet_name = _get_rfa_sheets_config(session)
    if not spreadsheet_id:
        return None
    try:
        from app.services.sheets_loader import load_from_sheets
    except ImportError:
        return None
    try:
        data_list, raw_columns, column_mapping = load_from_sheets(
            spreadsheet_id=spreadsheet_id,
            sheet_name_or_range=sheet_name,
        )
    except Exception:
        return None
    if not data_list:
        return None
    set_live_import(raw_columns, column_mapping, data_list)
    if session:
        try:
            from app.services.rfa_supabase import write_rfa_to_supabase
            write_rfa_to_supabase(session, data_list)
        except Exception as e:
            print(f"[RESOLVE] Erreur écriture rfa_data: {e}")
    import_data = get_live_import()
    if import_data:
        try:
            compute_aggregations(import_data)
        except Exception:
            pass
    return import_data


@router.get("/rfa-sheets/kpis")
async def get_rfa_kpis(import_id: str = "sheets_live", session: Session = Depends(get_session)):
    """KPIs rapides pour Nicolas (pas de calcul RFA complet — juste les CA agrégés)."""
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        return {"nb_clients": 0, "nb_groupes": 0, "ca_total": 0, "ca_by_supplier": {}}
    by_client = import_data.by_client or {}
    by_group  = import_data.by_group or {}
    ca_total = sum(c.get("grand_total", 0) for c in by_client.values())
    from app.core.fields import get_global_fields
    ca_by_supplier = {}
    for key in get_global_fields():
        label = key.replace("GLOBAL_", "")
        ca_by_supplier[label] = sum(
            c.get("global", {}).get(key, 0) for c in by_client.values()
        )
    top_clients = sorted(
        [{"id": k, "label": f"{k} - {v.get('nom_client','')}", "ca": v.get("grand_total", 0)}
         for k, v in by_client.items()],
        key=lambda x: -x["ca"]
    )[:10]
    return {
        "nb_clients":    len(by_client),
        "nb_groupes":    len(by_group),
        "ca_total":      ca_total,
        "ca_by_supplier": ca_by_supplier,
        "top_clients":   top_clients,
    }


@router.get("/rfa-sheets/config")
async def get_rfa_sheets_config(session: Session = Depends(get_session)):
    """Retourne la config de la feuille RFA connectée (admin + tous pour affichage)."""
    spreadsheet_id, sheet_name = _get_rfa_sheets_config(session)
    return {
        "spreadsheet_id": spreadsheet_id or "",
        "sheet_name": sheet_name or "",
        "configured": bool(spreadsheet_id),
    }


@router.put("/rfa-sheets/config")
async def set_rfa_sheets_config(
    body: Dict[str, Any] = Body(...),
    session: Session = Depends(get_session),
):
    """Enregistre la feuille RFA comme source (admin)."""
    spreadsheet_id = (body.get("spreadsheet_id") or "").strip()
    sheet_name = (body.get("sheet_name") or "").strip() or None
    if not spreadsheet_id:
        raise HTTPException(status_code=400, detail="spreadsheet_id requis")
    for key, value in [
        ("rfa_sheets_spreadsheet_id", spreadsheet_id),
        ("rfa_sheets_sheet_name", sheet_name or ""),
    ]:
        st = session.exec(select(AppSettings).where(AppSettings.key == key)).first()
        if st:
            st.value = value
            st.updated_at = datetime.now()
        else:
            session.add(AppSettings(key=key, value=value))
    session.commit()
    return {"spreadsheet_id": spreadsheet_id, "sheet_name": sheet_name or ""}


@router.post("/rfa-sheets/refresh", response_model=UploadResponse)
async def refresh_rfa_sheets(
    body: Optional[Dict[str, Any]] = Body(None),
    session: Session = Depends(get_session),
):
    """
    Met à jour les données RFA depuis la feuille Sheets.
    Body optionnel : { "spreadsheet_id": "...", "sheet_name": "..." }. Si fourni, utilise ces valeurs (et les enregistre en config).
    Sinon utilise la config déjà enregistrée ou .env.
    """
    try:
        from app.services.sheets_loader import load_from_sheets
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail="Google Sheets non configuré. Installez : pip install -r requirements-sheets.txt",
        ) from e
    spreadsheet_id = (body or {}).get("spreadsheet_id") or ""
    sheet_name_body = (body or {}).get("sheet_name") or ""
    if spreadsheet_id:
        spreadsheet_id = spreadsheet_id.strip()
        sheet_name_body = (sheet_name_body or "").strip() or None
    if not spreadsheet_id:
        spreadsheet_id, sheet_name_body = _get_rfa_sheets_config(session)
        sheet_name_body = sheet_name_body or None
    if not spreadsheet_id:
        raise HTTPException(
            status_code=400,
            detail="Aucune feuille RFA configurée. Définissez rfa_sheets_spreadsheet_id (paramètres ou .env).",
        )
    try:
        data, raw_columns, column_mapping = load_from_sheets(
            spreadsheet_id=spreadsheet_id,
            sheet_name_or_range=sheet_name_body,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur Google Sheets: {str(e)}")
    if not data:
        raise HTTPException(
            status_code=400,
            detail="Aucune donnée valide dans le Sheet.",
        )
    # Sauvegarde dans la table rfa_data Supabase (source de vérité pour les cold starts)
    try:
        from app.services.rfa_supabase import write_rfa_to_supabase
        nb = write_rfa_to_supabase(session, data)
        print(f"[REFRESH] {nb} lignes écrites dans rfa_data")
    except Exception as e:
        print(f"[REFRESH] Erreur écriture rfa_data: {e}")
        # Fallback : ancien cache JSON
        _save_live_cache(session, data, raw_columns, column_mapping)

    if body and body.get("spreadsheet_id"):
        for key, value in [
            ("rfa_sheets_spreadsheet_id", spreadsheet_id),
            ("rfa_sheets_sheet_name", sheet_name_body or ""),
        ]:
            st = session.exec(select(AppSettings).where(AppSettings.key == key)).first()
            if st:
                st.value = value
                st.updated_at = datetime.now()
            else:
                session.add(AppSettings(key=key, value=value))
        session.commit()
    import_data = get_live_import()
    if import_data:
        try:
            compute_aggregations(import_data)
        except Exception as agg_error:
            import traceback
            print(traceback.format_exc())
            raise HTTPException(
                status_code=500,
                detail=f"Erreur agrégation: {str(agg_error)}",
            )
        # Note: genie_full_analysis est trop lent pour Vercel (>10s) — pas mis en cache ici
    return UploadResponse(
        import_id=LIVE_IMPORT_ID,
        meta={"source": "google_sheets", "spreadsheet_id": spreadsheet_id, "nb_lignes": len(data)},
        nb_lignes=len(data),
        colonnes_brutes=raw_columns,
        colonnes_reconnues=column_mapping,
    )


@router.get("/rfa-sheets/current")
async def get_rfa_sheets_current(session: Session = Depends(get_session)):
    """
    Retourne l'import_id à utiliser quand la source est la feuille Sheets.
    Le frontend peut l'utiliser comme currentImportId pour que les utilisateurs voient les données sans importer.
    """
    _resolve_import_data(LIVE_IMPORT_ID, session)
    has_data = get_live_import() is not None
    return {"import_id": LIVE_IMPORT_ID, "has_data": has_data}


@router.get("/imports/{import_id}/clients", response_model=List[ClientSummary])
async def get_clients(import_id: str, session: Session = Depends(get_session)):
    """Liste des clients pour un import."""
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        raise HTTPException(status_code=404, detail="Import non trouvé")
    
    return get_clients_summary(import_data)


@router.get("/imports/{import_id}/client/{code_union}", response_model=ClientDetail)
async def get_client(import_id: str, code_union: str, session: Session = Depends(get_session)):
    """Détail d'un client."""
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        available_imports = list_imports()
        raise HTTPException(
            status_code=404, 
            detail=f"Import non trouve (ID: {import_id}). Les imports sont stockes en memoire et peuvent etre perdus apres un redemarrage du serveur. Imports disponibles: {len(available_imports)}"
        )
    
    try:
        return get_client_detail(import_data, code_union)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/imports/{import_id}/entities", response_model=List[EntitySummary])
async def get_entities(import_id: str, mode: str = "client", session: Session = Depends(get_session)):
    """
    Liste des entités (clients ou groupes) selon le mode.
    mode: "client" ou "group"
    """
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        available_imports = list_imports()
        raise HTTPException(
            status_code=404, 
            detail=f"Import non trouve (ID: {import_id}). Les imports sont stockes en memoire et peuvent etre perdus apres un redemarrage du serveur. Imports disponibles: {len(available_imports)}"
        )
    
    if mode not in ["client", "group"]:
        raise HTTPException(status_code=400, detail="mode doit être 'client' ou 'group'")
    
    # Calculer les agrégations si nécessaire
    if len(import_data.by_client) == 0 or len(import_data.by_group) == 0:
        try:
            compute_aggregations(import_data)
        except Exception as e:
            import traceback
            print(f"Erreur lors de l'agrégation: {e}")
            print(traceback.format_exc())
            raise HTTPException(
                status_code=500,
                detail=f"Erreur lors de l'agrégation: {str(e)}"
            )
    
    entities = []

    if mode == "client":
        for code_union, data in import_data.by_client.items():
            nom = data.get("nom_client") or ""
            label = f"{code_union} - {nom}" if nom else code_union
            entities.append(EntitySummary(
                id=code_union,
                label=label,
                groupe_client=data.get("groupe_client"),
                global_total=data["global_total"],
                tri_total=data["tri_total"],
                grand_total=data["grand_total"],
                rfa_total=None,  # calculé à la demande (fiche client)
            ))
        entities.sort(key=lambda x: x.label)
    
    else:  # mode == "group"
        for groupe, data in import_data.by_group.items():
            entities.append(EntitySummary(
                id=groupe,
                label=groupe,
                nb_comptes=data["nb_comptes"],
                global_total=data["global_total"],
                tri_total=data["tri_total"],
                grand_total=data["grand_total"],
                rfa_total=None,  # calculé à la demande (fiche groupe)
            ))
        entities.sort(key=lambda x: x.label)
    
    return entities


@router.get("/imports/{import_id}/entity", response_model=EntityDetailWithRfa)
async def get_entity(
    import_id: str,
    mode: str,
    id: str,
    contract_id: Optional[int] = Query(None, description="ID du contrat pour simulation"),
    session: Session = Depends(get_session),
):
    """
    Détail d'une entité (client ou groupe) avec calcul RFA.
    mode: "client" ou "group"
    id: code_union (si mode=client) ou groupe_client (si mode=group)
    contract_id: ID du contrat à utiliser (optionnel, pour simulation)
    """
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        available_imports = list_imports()
        raise HTTPException(
            status_code=404, 
            detail=f"Import non trouve (ID: {import_id}). Les imports sont stockes en memoire et peuvent etre perdus apres un redemarrage du serveur. Imports disponibles: {len(available_imports)}"
        )
    
    if mode not in ["client", "group"]:
        raise HTTPException(status_code=400, detail="mode doit etre 'client' ou 'group'")
    
    try:
        return get_entity_detail_with_rfa(import_data, mode, id, contract_id=contract_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/imports/{import_id}/union")
async def get_union_entity(import_id: str, session: Session = Depends(get_session)):
    """
    Détail Union (agrégation globale tous clients) avec calcul RFA.
    """
    print(f"\n[UNION ENDPOINT] Requete recue pour import_id: {import_id}")
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        available_imports = list_imports()
        print(f"[UNION ENDPOINT] Import non trouve : {import_id}")
        raise HTTPException(
            status_code=404,
            detail=f"Import non trouve (ID: {import_id}). Imports disponibles: {len(available_imports)}"
        )
    
    print(f"[UNION ENDPOINT] Import trouve, calcul en cours...")
    try:
        from app.services.compute import get_union_detail_with_rfa
        result = get_union_detail_with_rfa(import_data)
        print(f"[UNION ENDPOINT] Calcul termine avec succes")
        return result
    except Exception as e:
        import traceback
        import sys
        sys.stderr.write("\n" + "="*80 + "\n")
        sys.stderr.write("[UNION ENDPOINT] ERREUR DETAILLEE :\n")
        sys.stderr.write(traceback.format_exc())
        sys.stderr.write("="*80 + "\n")
        raise HTTPException(status_code=500, detail=f"Erreur calcul Union: {str(e)}")


@router.get("/imports/{import_id}/union/export-excel")
async def export_union_excel(import_id: str, session: Session = Depends(get_session)):
    """Export Excel des donnees Union RFA."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        raise HTTPException(status_code=404, detail="Import non trouve")
    
    try:
        from app.services.compute import get_union_detail_with_rfa
        result = get_union_detail_with_rfa(import_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur calcul: {str(e)}")
    
    wb = openpyxl.Workbook()
    
    # --- Feuille 1 : Synthese ---
    ws = wb.active
    ws.title = "Synthese Union"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    money_fmt = '#,##0.00 "EUR"'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Titre
    ws.merge_cells('A1:G1')
    ws['A1'] = "RAPPORT RFA UNION - GROUPEMENT"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A2'] = f"Import ID: {import_id}"
    ws['A2'].font = Font(italic=True, color="808080")
    ws['A3'] = f"Nombre de comptes: {result.nb_comptes}"
    
    # En-tetes RFA Globale
    row = 5
    headers = ["Fournisseur", "CA Global", "Taux RFA", "Montant RFA", "Taux Bonus", "Montant Bonus", "Total Ligne"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Donnees globales
    row = 6
    rfa_data = result.rfa
    if hasattr(rfa_data, 'global_items'):
        global_items = rfa_data.global_items
    elif hasattr(rfa_data, 'dict'):
        d = rfa_data.dict(by_alias=True) if hasattr(rfa_data, 'dict') else rfa_data.model_dump(by_alias=True)
        global_items = d.get('global', {})
    else:
        global_items = {}
    
    total_rfa_global = 0
    for key, item in (global_items.items() if isinstance(global_items, dict) else []):
        if isinstance(item, dict):
            ca = item.get('ca', 0)
            rfa_rate = item.get('rfa', {}).get('rate', 0)
            rfa_value = item.get('rfa', {}).get('value', 0)
            bonus_rate = item.get('bonus', {}).get('rate', 0)
            bonus_value = item.get('bonus', {}).get('value', 0)
            label = item.get('label', key)
        else:
            ca = item.ca
            rfa_rate = item.rfa.rate if item.rfa else 0
            rfa_value = item.rfa.value if item.rfa else 0
            bonus_rate = item.bonus.rate if item.bonus else 0
            bonus_value = item.bonus.value if item.bonus else 0
            label = item.label
        
        total_ligne = rfa_value + bonus_value
        total_rfa_global += total_ligne
        
        if ca == 0 and total_ligne == 0:
            continue
        
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=ca).number_format = money_fmt
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=rfa_rate).number_format = pct_fmt
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=rfa_value).number_format = money_fmt
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5, value=bonus_rate).number_format = pct_fmt
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6, value=bonus_value).number_format = money_fmt
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=7, value=total_ligne).number_format = money_fmt
        ws.cell(row=row, column=7).border = thin_border
        row += 1
    
    # Sous-total Global
    ws.cell(row=row, column=1, value="TOTAL RFA GLOBALE").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=7, value=total_rfa_global).number_format = money_fmt
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=7).fill = subtotal_fill
    ws.cell(row=row, column=7).border = thin_border
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = subtotal_fill
        ws.cell(row=row, column=c).border = thin_border
    row += 2
    
    # En-tetes Tri-partites
    tri_headers = ["Tri-partite", "CA", "Seuil Min", "Taux", "Montant RFA"]
    for col, h in enumerate(tri_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1
    
    # Donnees tri-partites
    if hasattr(rfa_data, 'tri_items'):
        tri_items = rfa_data.tri_items
    elif hasattr(rfa_data, 'dict'):
        d = rfa_data.dict(by_alias=True) if hasattr(rfa_data, 'dict') else rfa_data.model_dump(by_alias=True)
        tri_items = d.get('tri', {})
    else:
        tri_items = {}
    
    total_rfa_tri = 0
    for key, item in (tri_items.items() if isinstance(tri_items, dict) else []):
        if isinstance(item, dict):
            ca = item.get('ca', 0)
            rate = item.get('rate', 0)
            value = item.get('value', 0)
            label = item.get('label', key)
            sel_min = item.get('selected_min', None)
        else:
            ca = item.ca
            rate = item.rate
            value = item.value
            label = item.label
            sel_min = item.selected_min
        
        if ca == 0 and value == 0:
            continue
        
        total_rfa_tri += value
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=ca).number_format = money_fmt
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3, value=sel_min or 0).number_format = money_fmt
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=rate).number_format = pct_fmt
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5, value=value).number_format = money_fmt
        ws.cell(row=row, column=5).border = thin_border
        row += 1
    
    # Sous-total Tri
    ws.cell(row=row, column=1, value="TOTAL TRI-PARTITES").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=5, value=total_rfa_tri).number_format = money_fmt
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = subtotal_fill
    ws.cell(row=row, column=5).border = thin_border
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = subtotal_fill
        ws.cell(row=row, column=c).border = thin_border
    row += 2
    
    # Grand Total
    grand_total = total_rfa_global + total_rfa_tri
    ca_total = result.ca.get('totals', {}).get('grand_total', 0) if isinstance(result.ca, dict) else 0
    ws.cell(row=row, column=1, value="GRAND TOTAL RFA UNION").font = Font(bold=True, size=12, color="1F4E79")
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=ca_total).number_format = money_fmt
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = thin_border
    taux_global = grand_total / ca_total if ca_total > 0 else 0
    ws.cell(row=row, column=3, value=taux_global).number_format = pct_fmt
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=4, value=grand_total).number_format = money_fmt
    ws.cell(row=row, column=4).font = Font(bold=True, size=12, color="006600")
    ws.cell(row=row, column=4).border = thin_border
    
    # Ajuster largeurs
    for col_letter, width in [('A', 30), ('B', 18), ('C', 14), ('D', 18), ('E', 14), ('F', 18), ('G', 18)]:
        ws.column_dimensions[col_letter].width = width
    
    # Sauvegarder
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="RFA_Union_{import_id[:8]}.xlsx"'}
    )


@router.get("/imports/{import_id}/union/export-pdf")
async def export_union_pdf(import_id: str, session: Session = Depends(get_session)):
    """Export PDF du rapport Union RFA."""
    try:
        from xhtml2pdf import pisa
    except ImportError:
        raise HTTPException(status_code=503, detail="Export PDF non disponible dans cet environnement.")
    from io import BytesIO
    
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        raise HTTPException(status_code=404, detail="Import non trouve")
    
    try:
        from app.services.compute import get_union_detail_with_rfa
        result = get_union_detail_with_rfa(import_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur calcul: {str(e)}")
    
    def fmt_amount(v):
        try:
            return f"{v:,.0f} EUR".replace(",", " ")
        except:
            return "0 EUR"
    
    def fmt_pct(v):
        try:
            return f"{v*100:.2f} %"
        except:
            return "0 %"
    
    # Extraire les donnees
    rfa_data = result.rfa
    if hasattr(rfa_data, 'model_dump'):
        rfa_dict = rfa_data.model_dump(by_alias=True)
    elif hasattr(rfa_data, 'dict'):
        rfa_dict = rfa_data.dict(by_alias=True)
    else:
        rfa_dict = {"global": {}, "tri": {}, "totals": {}}
    
    global_items = rfa_dict.get('global', {})
    tri_items = rfa_dict.get('tri', {})
    totals = rfa_dict.get('totals', {})
    ca_data = result.ca if isinstance(result.ca, dict) else {}
    ca_total = ca_data.get('totals', {}).get('grand_total', 0)
    grand_total_rfa = totals.get('grand_total', 0)
    taux_global = grand_total_rfa / ca_total if ca_total > 0 else 0
    
    # Construire les lignes globales
    global_rows_html = ""
    for key, item in global_items.items():
        ca = item.get('ca', 0)
        rfa_v = item.get('rfa', {}).get('value', 0)
        rfa_r = item.get('rfa', {}).get('rate', 0)
        bonus_v = item.get('bonus', {}).get('value', 0)
        bonus_r = item.get('bonus', {}).get('rate', 0)
        total = rfa_v + bonus_v
        label = item.get('label', key)
        if ca == 0 and total == 0:
            continue
        triggered = item.get('triggered', False)
        color = "#e8f5e9" if triggered else "#fff"
        global_rows_html += f"""
        <tr style="background:{color}">
            <td style="padding:8px;border:1px solid #ddd;font-weight:600">{label}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_amount(ca)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_pct(rfa_r)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_amount(rfa_v)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_pct(bonus_r)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_amount(bonus_v)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right;font-weight:700">{fmt_amount(total)}</td>
        </tr>"""
    
    # Construire les lignes tri-partites
    tri_rows_html = ""
    for key, item in tri_items.items():
        ca = item.get('ca', 0)
        rate = item.get('rate', 0)
        value = item.get('value', 0)
        label = item.get('label', key)
        if ca == 0 and value == 0:
            continue
        triggered = item.get('triggered', False)
        color = "#e8f5e9" if triggered else "#fff"
        tri_rows_html += f"""
        <tr style="background:{color}">
            <td style="padding:8px;border:1px solid #ddd;font-weight:600">{label}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_amount(ca)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_pct(rate)}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right;font-weight:700">{fmt_amount(value)}</td>
        </tr>"""
    
    html = f"""
    <html>
    <head>
        <meta charset="utf-8" />
        <style>
            body {{ font-family: Helvetica, Arial, sans-serif; font-size: 10px; color: #333; margin: 20px; }}
            h1 {{ color: #1F4E79; font-size: 20px; margin-bottom: 5px; }}
            h2 {{ color: #1F4E79; font-size: 14px; margin-top: 25px; margin-bottom: 8px; border-bottom: 2px solid #1F4E79; padding-bottom: 4px; }}
            .kpi-box {{ display: inline-block; width: 30%; padding: 12px; margin: 5px; background: #f0f4f8; border-radius: 8px; text-align: center; }}
            .kpi-value {{ font-size: 18px; font-weight: bold; color: #1F4E79; }}
            .kpi-label {{ font-size: 9px; color: #666; margin-top: 2px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
            th {{ background: #1F4E79; color: white; padding: 8px; text-align: center; font-size: 9px; border: 1px solid #1F4E79; }}
            .th-green {{ background: #548235; border-color: #548235; }}
            .total-row {{ background: #D6E4F0; font-weight: bold; }}
            .grand-total {{ background: #1F4E79; color: white; font-weight: bold; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h1>RAPPORT RFA UNION - GROUPEMENT</h1>
        <p style="color:#888;font-size:9px">Import: {import_id[:8]} | Comptes: {result.nb_comptes} | Contrats: {result.contract_applied.get('name', '') if isinstance(result.contract_applied, dict) else ''}</p>
        
        <div style="margin: 15px 0">
            <div class="kpi-box">
                <div class="kpi-value">{fmt_amount(ca_total)}</div>
                <div class="kpi-label">CA TOTAL UNION</div>
            </div>
            <div class="kpi-box">
                <div class="kpi-value" style="color:#006600">{fmt_amount(grand_total_rfa)}</div>
                <div class="kpi-label">RFA TOTALE</div>
            </div>
            <div class="kpi-box">
                <div class="kpi-value" style="color:#4472C4">{fmt_pct(taux_global)}</div>
                <div class="kpi-label">TAUX MOYEN</div>
            </div>
        </div>
        
        <h2>RFA Globales par Fournisseur</h2>
        <table>
            <tr>
                <th>Fournisseur</th><th>CA Global</th><th>Taux RFA</th><th>Montant RFA</th>
                <th>Taux Bonus</th><th>Montant Bonus</th><th>Total</th>
            </tr>
            {global_rows_html}
            <tr class="total-row">
                <td style="padding:8px;border:1px solid #ddd">TOTAL GLOBALES</td>
                <td colspan="5" style="border:1px solid #ddd"></td>
                <td style="padding:8px;border:1px solid #ddd;text-align:right;font-weight:700">{fmt_amount(totals.get('global_total', 0))}</td>
            </tr>
        </table>
        
        <h2>RFA Tri-partites</h2>
        <table>
            <tr>
                <th class="th-green">Tri-partite</th><th class="th-green">CA</th>
                <th class="th-green">Taux</th><th class="th-green">Montant RFA</th>
            </tr>
            {tri_rows_html}
            <tr class="total-row">
                <td style="padding:8px;border:1px solid #ddd">TOTAL TRI-PARTITES</td>
                <td colspan="2" style="border:1px solid #ddd"></td>
                <td style="padding:8px;border:1px solid #ddd;text-align:right;font-weight:700">{fmt_amount(totals.get('tri_total', 0))}</td>
            </tr>
        </table>
        
        <div style="margin-top:20px;padding:15px;background:#1F4E79;color:white;border-radius:8px;text-align:center">
            <div style="font-size:10px">GRAND TOTAL RFA UNION</div>
            <div style="font-size:22px;font-weight:bold;margin:5px 0">{fmt_amount(grand_total_rfa)}</div>
            <div style="font-size:10px">sur un CA de {fmt_amount(ca_total)} soit un taux global de {fmt_pct(taux_global)}</div>
        </div>
    </body>
    </html>
    """
    
    output = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=output)
    
    if pisa_status.err:
        raise HTTPException(status_code=500, detail="Erreur generation PDF")
    
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="RFA_Union_{import_id[:8]}.pdf"'}
    )


@router.get("/imports/{import_id}/recap", response_model=RecapGlobalRfa)
async def get_global_recap(
    import_id: str,
    dissolved_groups: Optional[str] = Query(None, description="Liste des groupes dissous (séparés par des virgules)"),
    session: Session = Depends(get_session),
):
    """
    Récapitulatif global RFA sans double comptage.
    
    Compte :
    - Les clients qui n'ont PAS de groupe_client
    - Les groupes (qui incluent déjà leurs clients)
    - Les clients des groupes dissous (traités individuellement)
    
    Args:
        dissolved_groups: Liste des noms de groupes à traiter individuellement (ex: "INDEPENDANT UNION,GROUPE ABC")
    """
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        available_imports = list_imports()
        raise HTTPException(
            status_code=404, 
            detail=f"Import non trouve (ID: {import_id}). Les imports sont stockes en memoire et peuvent etre perdus apres un redemarrage du serveur. Imports disponibles: {len(available_imports)}"
        )
    
    # Parser les groupes dissous
    dissolved_set = set()
    if dissolved_groups:
        dissolved_set = {g.strip().upper() for g in dissolved_groups.split(",") if g.strip()}

    try:
        import json as _json
        from app.services.contract_resolver import BatchContractResolver
        from app.models import ContractRule, ContractOverride
        import app.services.contract_resolver as _resolver_mod
        import app.services.rfa_calculator as _rfa_calc

        # ── Batch 1 : résolution des contrats (3 requêtes) ──────────────
        _batch = BatchContractResolver()
        _orig_resolve = _resolver_mod.resolve_contract
        _resolver_mod.resolve_contract = lambda code_union=None, groupe_client=None: _batch.resolve(code_union, groupe_client)

        # ── Batch 2 : toutes les règles de contrats (1 requête) ─────────
        _all_rules     = session.exec(select(ContractRule)).all()
        _all_overrides = session.exec(select(ContractOverride).where(ContractOverride.is_active == True)).all()

        _rules_cache = {}
        for _r in _all_rules:
            _rules_cache.setdefault(_r.contract_id, {})[_r.key] = _r

        # ── Batch 3 : tous les overrides (1 requête) ────────────────────
        _ov_cache = {}
        for _ov in _all_overrides:
            _tt = _ov.target_type.value if hasattr(_ov.target_type, 'value') else str(_ov.target_type)
            _k  = (_tt, (_ov.target_value or "").strip().upper())
            _ov_cache.setdefault(_k, {}).setdefault(_ov.field_key, {})
            try:
                _ov_cache[_k][_ov.field_key][_ov.tier_type.value if hasattr(_ov.tier_type,'value') else str(_ov.tier_type)] = _json.loads(_ov.custom_tiers)
            except Exception:
                pass

        _orig_rules    = _rfa_calc.load_contract_rules
        _orig_overrides = _rfa_calc.load_entity_overrides
        _rfa_calc.load_contract_rules    = lambda c: _rules_cache.get(c.id, {})
        _rfa_calc.load_entity_overrides  = lambda tt, tv: _ov_cache.get(
            (str(tt) if not hasattr(tt,'value') else tt.value, (tv or "").strip().upper()), {}
        )

        try:
            result = get_global_recap_rfa(import_data, dissolved_groups=dissolved_set)
        finally:
            _resolver_mod.resolve_contract   = _orig_resolve
            _rfa_calc.load_contract_rules    = _orig_rules
            _rfa_calc.load_entity_overrides  = _orig_overrides

        return result
    except Exception as e:
        import traceback
        print(f"Erreur lors du calcul du récapitulatif: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur lors du calcul du récapitulatif: {str(e)}")


@router.get("/imports/{import_id}/entity/pdf")
async def get_entity_pdf(
    import_id: str,
    mode: str,
    id: str,
    contract_id: Optional[int] = Query(None, description="ID du contrat pour simulation"),
    session: Session = Depends(get_session),
):
    """
    Génère un PDF pour une entité (client ou groupe) avec calcul RFA.
    mode: "client" ou "group"
    id: code_union (si mode=client) ou groupe_client (si mode=group)
    """
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        available_imports = list_imports()
        raise HTTPException(
            status_code=404,
            detail=f"Import non trouve (ID: {import_id}). Les imports sont stockes en memoire et peuvent etre perdus apres un redemarrage du serveur. Imports disponibles: {len(available_imports)}"
        )
    
    if mode not in ["client", "group"]:
        raise HTTPException(status_code=400, detail="mode doit etre 'client' ou 'group'")
    
    try:
        pdf_buffer = generate_pdf_report(import_id, mode, id, contract_id=contract_id)
        
        # Déterminer le nom du fichier
        entity_label = id.replace(" ", "_")
        filename = f"RFA_{entity_label}_{mode}.pdf"
        
        return Response(
            content=pdf_buffer.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        import traceback
        print(f"Erreur lors de la generation du PDF: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur lors de la generation du PDF: {str(e)}")


# ==================== ENDPOINTS CONTRATS ====================

@router.get("/contracts")
async def list_contracts(session: Session = Depends(get_session)):
    """Liste tous les contrats."""
    statement = select(Contract).order_by(Contract.is_default.desc(), Contract.name)
    contracts = session.exec(statement).all()
    return contracts


@router.post("/contracts")
async def create_contract(contract: Contract, session: Session = Depends(get_session)):
    """Crée un nouveau contrat."""
    from app.models import ContractScope
    # Un seul contrat par défaut par scope (Adhérent / Union)
    if contract.is_default:
        statement = select(Contract).where(
            Contract.is_default == True,
            Contract.scope == contract.scope,
        )
        existing_default = session.exec(statement).first()
        if existing_default:
            existing_default.is_default = False
            session.add(existing_default)
    
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.get("/contracts/{contract_id}")
async def get_contract(contract_id: int, session: Session = Depends(get_session)):
    """Récupère un contrat par ID."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    return contract


@router.put("/contracts/{contract_id}")
async def update_contract(contract_id: int, contract_update: Contract, session: Session = Depends(get_session)):
    """Met à jour un contrat."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    from app.models import ContractScope
    # Un seul contrat par défaut par scope
    if contract_update.is_default and not contract.is_default:
        statement = select(Contract).where(
            Contract.is_default == True,
            Contract.scope == contract_update.scope,
        )
        existing_default = session.exec(statement).first()
        if existing_default:
            existing_default.is_default = False
            session.add(existing_default)
    
    contract.name = contract_update.name
    contract.description = contract_update.description
    # Persister le scope (normaliser string -> enum au cas où le client envoie une chaîne)
    scope_val = getattr(contract_update, "scope", None)
    if scope_val is not None:
        if isinstance(scope_val, str):
            try:
                contract.scope = ContractScope[scope_val.upper()]
            except KeyError:
                pass  # garder la valeur existante
        else:
            contract.scope = scope_val
    contract.marketing_rules = contract_update.marketing_rules
    contract.is_default = contract_update.is_default
    contract.is_active = contract_update.is_active
    contract.updated_at = datetime.now()
    
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.put("/contracts/{contract_id}/set-default")
async def set_default_contract(contract_id: int, session: Session = Depends(get_session)):
    """Définit un contrat comme défaut (un défaut par scope : un pour Adhérent, un pour Union/DAF)."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    # Désactiver les autres par défaut du même scope uniquement
    from app.models import ContractScope
    statement = select(Contract).where(
        Contract.is_default == True,
        Contract.scope == contract.scope,
    )
    existing_defaults = session.exec(statement).all()
    for c in existing_defaults:
        c.is_default = False
        session.add(c)
    
    contract.is_default = True
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.put("/contracts/{contract_id}/toggle-active")
async def toggle_active_contract(contract_id: int, session: Session = Depends(get_session)):
    """Active/désactive un contrat."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    contract.is_active = not contract.is_active
    contract.updated_at = datetime.now()
    
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.post("/contracts/{contract_id}/duplicate")
async def duplicate_contract(contract_id: int, session: Session = Depends(get_session)):
    """Duplique un contrat avec toutes ses règles."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    # Créer le nouveau contrat
    new_contract = Contract(
        name=f"{contract.name} (copie)",
        description=contract.description,
        is_default=False,
        is_active=True
    )
    session.add(new_contract)
    session.commit()
    session.refresh(new_contract)
    
    # Dupliquer les règles
    statement = select(ContractRule).where(ContractRule.contract_id == contract.id)
    rules = session.exec(statement).all()
    for rule in rules:
        new_rule = ContractRule(
            contract_id=new_contract.id,
            key=rule.key,
            scope=rule.scope,
            label=rule.label,
            tiers_rfa=rule.tiers_rfa,
            tiers_bonus=rule.tiers_bonus,
            tiers=rule.tiers
        )
        session.add(new_rule)
    
    session.commit()
    session.refresh(new_contract)
    return new_contract


@router.delete("/contracts/{contract_id}")
async def delete_contract(contract_id: int, session: Session = Depends(get_session)):
    """Supprime un contrat et toutes ses règles et affectations."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    # Vérifier si c'est le contrat par défaut
    if contract.is_default:
        raise HTTPException(
            status_code=400,
            detail="Impossible de supprimer le contrat par défaut. Définissez un autre contrat par défaut d'abord."
        )
    
    # Supprimer les affectations associées
    statement_assignments = select(ContractAssignment).where(
        ContractAssignment.contract_id == contract_id
    )
    assignments = session.exec(statement_assignments).all()
    for assignment in assignments:
        session.delete(assignment)
    
    # Supprimer les règles
    statement_rules = select(ContractRule).where(
        ContractRule.contract_id == contract_id
    )
    rules = session.exec(statement_rules).all()
    for rule in rules:
        session.delete(rule)
    
    # Supprimer le contrat
    session.delete(contract)
    session.commit()
    
    return {"message": f"Contrat '{contract.name}' supprimé avec succès"}


# ==================== ENDPOINTS RÈGLES ====================

@router.get("/contracts/available-tri-fields")
async def get_available_tri_fields():
    """Retourne la liste de toutes les clés tri-partites connues (key + label). Permet à l'éditeur de contrat d'afficher toutes les colonnes, y compris celles sans règle."""
    from app.core.fields import get_tri_fields, get_field_by_key
    result = []
    for key in get_tri_fields():
        _, label = get_field_by_key(key)
        result.append({"key": key, "label": label})
    return result


@router.get("/contracts/{contract_id}/rules")
async def get_contract_rules(contract_id: int, session: Session = Depends(get_session)):
    """Récupère toutes les règles d'un contrat."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    statement = select(ContractRule).where(ContractRule.contract_id == contract_id)
    rules = session.exec(statement).all()
    return rules


@router.put("/contracts/{contract_id}/rules/{rule_id}")
async def update_contract_rule(
    contract_id: int,
    rule_id: int,
    rule_update: ContractRule,
    session: Session = Depends(get_session)
):
    """Met à jour une règle de contrat."""
    rule = session.get(ContractRule, rule_id)
    if not rule or rule.contract_id != contract_id:
        raise HTTPException(status_code=404, detail="Règle non trouvée")
    
    rule.label = rule_update.label
    rule.tiers_rfa = rule_update.tiers_rfa
    rule.tiers_bonus = rule_update.tiers_bonus
    rule.tiers = rule_update.tiers
    rule.updated_at = datetime.now()
    
    session.add(rule)
    session.commit()
    session.refresh(rule)
    return rule


@router.post("/contracts/{contract_id}/rules")
async def create_contract_rule(
    contract_id: int,
    body: dict,
    session: Session = Depends(get_session)
):
    """Crée une nouvelle règle de contrat (ex: pour une tri-partite non encore configurée)."""
    contract = session.get(Contract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    key = body.get("key")
    label = body.get("label", key or "")
    scope_str = body.get("scope", "TRI")
    try:
        scope = RuleScope[scope_str]
    except KeyError:
        scope = RuleScope.TRI
    # Vérifier qu'une règle avec cette clé n'existe pas déjà
    statement = select(ContractRule).where(
        ContractRule.contract_id == contract_id,
        ContractRule.key == key
    )
    if session.exec(statement).first():
        raise HTTPException(status_code=400, detail=f"Une règle pour '{key}' existe déjà")
    from app.core.fields import get_tri_fields, get_field_by_key
    if scope == RuleScope.TRI and key not in get_tri_fields():
        raise HTTPException(status_code=400, detail=f"Clé tri-partite inconnue: {key}")
    if scope == RuleScope.TRI:
        _, label = get_field_by_key(key)
    tiers = body.get("tiers")
    if tiers is None and scope == RuleScope.TRI:
        tiers = []
    tiers_rfa = body.get("tiers_rfa")
    tiers_bonus = body.get("tiers_bonus")
    rule = ContractRule(
        contract_id=contract_id,
        key=key,
        scope=scope,
        label=label,
        tiers=json.dumps(tiers) if isinstance(tiers, list) else tiers,
        tiers_rfa=json.dumps(tiers_rfa) if isinstance(tiers_rfa, list) else tiers_rfa,
        tiers_bonus=json.dumps(tiers_bonus) if isinstance(tiers_bonus, list) else tiers_bonus,
    )
    session.add(rule)
    session.commit()
    session.refresh(rule)
    return rule


# ==================== ENDPOINTS AFFECTATIONS ====================

@router.get("/assignments")
async def list_assignments(session: Session = Depends(get_session)):
    """Liste toutes les affectations."""
    statement = select(ContractAssignment).order_by(ContractAssignment.target_type, ContractAssignment.target_value)
    assignments = session.exec(statement).all()
    return assignments


@router.post("/assignments")
async def create_assignment(assignment: ContractAssignment, session: Session = Depends(get_session)):
    """Crée une nouvelle affectation."""
    # Normaliser la valeur (trim + uppercase pour cohérence)
    normalized_value = assignment.target_value.strip().upper()
    
    # Vérifier unicité (target_type + target_value normalisé)
    statement = select(ContractAssignment).where(
        ContractAssignment.target_type == assignment.target_type
    )
    all_assignments = session.exec(statement).all()
    for existing in all_assignments:
        if existing.target_value.strip().upper() == normalized_value:
            raise HTTPException(
                status_code=400,
                detail=f"Une affectation existe déjà pour {assignment.target_type.value}={normalized_value}"
            )
    
    # Stocker la valeur normalisée
    assignment.target_value = normalized_value
    
    # Définir la priorité automatiquement
    if assignment.target_type == TargetType.CODE_UNION:
        assignment.priority = 100
    else:
        assignment.priority = 50
    
    session.add(assignment)
    session.commit()
    session.refresh(assignment)
    return assignment


@router.delete("/assignments/{assignment_id}")
async def delete_assignment(assignment_id: int, session: Session = Depends(get_session)):
    """Supprime une affectation."""
    assignment = session.get(ContractAssignment, assignment_id)
    if not assignment:
        raise HTTPException(status_code=404, detail="Affectation non trouvée")
    
    session.delete(assignment)
    session.commit()
    return {"message": "Affectation supprimée"}


# ==================== ENDPOINTS OVERRIDES (Taux personnalises par client) ====================

@router.get("/overrides")
async def list_overrides(
    target_type: Optional[str] = Query(None, description="Filtrer par type (CODE_UNION ou GROUPE_CLIENT)"),
    target_value: Optional[str] = Query(None, description="Filtrer par valeur cible"),
    session: Session = Depends(get_session)
):
    """Liste tous les overrides, optionnellement filtres par target_type et target_value."""
    statement = select(ContractOverride)
    
    if target_type:
        try:
            target_type_enum = TargetType(target_type)
            statement = statement.where(ContractOverride.target_type == target_type_enum)
        except ValueError:
            pass  # Type invalide, ignorer le filtre
    
    if target_value:
        normalized = target_value.strip().upper()
        statement = statement.where(ContractOverride.target_value == normalized)
    
    statement = statement.order_by(ContractOverride.target_type, ContractOverride.target_value, ContractOverride.field_key)
    
    overrides = session.exec(statement).all()
    return overrides


@router.get("/overrides/entity/{target_type}/{target_value}")
async def get_entity_overrides(target_type: str, target_value: str, session: Session = Depends(get_session)):
    """Recupere tous les overrides pour une entite (client ou groupe)."""
    normalized = target_value.strip().upper()
    
    try:
        target_type_enum = TargetType(target_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"target_type invalide: {target_type}. Utiliser CODE_UNION ou GROUPE_CLIENT")
    
    statement = select(ContractOverride).where(
        ContractOverride.target_type == target_type_enum,
        ContractOverride.target_value == normalized
    ).order_by(ContractOverride.field_key)
    
    overrides = session.exec(statement).all()
    return overrides


@router.post("/overrides")
async def create_override(override: ContractOverride, session: Session = Depends(get_session)):
    """Cree un nouvel override de taux pour un client ou groupe."""
    # Normaliser la valeur cible
    normalized_value = override.target_value.strip().upper()
    
    # Verifier si un override existe deja pour cette combinaison
    statement = select(ContractOverride).where(
        ContractOverride.target_type == override.target_type,
        ContractOverride.target_value == normalized_value,
        ContractOverride.field_key == override.field_key,
        ContractOverride.tier_type == override.tier_type
    )
    existing = session.exec(statement).first()
    
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Un override existe deja pour {override.target_type.value}/{normalized_value} / {override.field_key} / {override.tier_type.value}. Utilisez PUT pour le mettre a jour."
        )
    
    # Valider le JSON des tiers
    try:
        import json
        tiers = json.loads(override.custom_tiers)
        if not isinstance(tiers, list):
            raise ValueError("custom_tiers doit etre un tableau")
        for tier in tiers:
            if "min" not in tier or "rate" not in tier:
                raise ValueError("Chaque tier doit avoir 'min' et 'rate'")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="custom_tiers doit etre un JSON valide")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # Creer l'override
    override.target_value = normalized_value
    session.add(override)
    session.commit()
    session.refresh(override)
    return override


@router.put("/overrides/{override_id}")
async def update_override(
    override_id: int,
    override_update: ContractOverride,
    session: Session = Depends(get_session)
):
    """Met a jour un override existant."""
    override = session.get(ContractOverride, override_id)
    if not override:
        raise HTTPException(status_code=404, detail="Override non trouve")
    
    # Valider le JSON des tiers
    try:
        import json
        tiers = json.loads(override_update.custom_tiers)
        if not isinstance(tiers, list):
            raise ValueError("custom_tiers doit etre un tableau")
        for tier in tiers:
            if "min" not in tier or "rate" not in tier:
                raise ValueError("Chaque tier doit avoir 'min' et 'rate'")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="custom_tiers doit etre un JSON valide")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # Mettre a jour
    override.custom_tiers = override_update.custom_tiers
    override.is_active = override_update.is_active
    override.updated_at = datetime.now()
    
    session.add(override)
    session.commit()
    session.refresh(override)
    return override


@router.delete("/overrides/{override_id}")
async def delete_override(override_id: int, session: Session = Depends(get_session)):
    """Supprime un override."""
    override = session.get(ContractOverride, override_id)
    if not override:
        raise HTTPException(status_code=404, detail="Override non trouve")
    
    session.delete(override)
    session.commit()
    return {"message": "Override supprime avec succes"}


@router.delete("/overrides/entity/{target_type}/{target_value}")
async def delete_all_entity_overrides(target_type: str, target_value: str, session: Session = Depends(get_session)):
    """Supprime tous les overrides d'une entite (client ou groupe)."""
    normalized = target_value.strip().upper()
    
    try:
        target_type_enum = TargetType(target_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"target_type invalide: {target_type}. Utiliser CODE_UNION ou GROUPE_CLIENT")
    
    statement = select(ContractOverride).where(
        ContractOverride.target_type == target_type_enum,
        ContractOverride.target_value == normalized
    )
    overrides = session.exec(statement).all()
    
    count = len(overrides)
    for override in overrides:
        session.delete(override)
    
    session.commit()
    return {"message": f"{count} override(s) supprime(s) pour {target_type}/{normalized}"}


# ==================== ENDPOINTS PUBLICITES ====================

@router.get("/ads", response_model=List[AdResponse])
async def list_ads(
    active_only: bool = Query(True, description="Ne retourner que les annonces actives et dans la periode"),
    session: Session = Depends(get_session)
):
    """Liste des annonces (logos/promo)."""
    statement = select(Ad)
    if active_only:
        now = datetime.now()
        statement = statement.where(
            Ad.is_active == True,
            (Ad.start_at == None) | (Ad.start_at <= now),
            (Ad.end_at == None) | (Ad.end_at >= now)
        )
    statement = statement.order_by(Ad.sort_order, Ad.created_at.desc())
    ads = session.exec(statement).all()
    return ads


@router.post("/ads", response_model=AdResponse)
async def create_ad(ad: AdCreate, session: Session = Depends(get_session)):
    """Cree une annonce."""
    ad_model = Ad(**ad.dict())
    ad_model.created_at = datetime.now()
    ad_model.updated_at = datetime.now()
    session.add(ad_model)
    session.commit()
    session.refresh(ad_model)
    return ad_model


@router.put("/ads/{ad_id}", response_model=AdResponse)
async def update_ad(ad_id: int, ad_update: AdUpdate, session: Session = Depends(get_session)):
    """Met a jour une annonce."""
    ad = session.get(Ad, ad_id)
    if not ad:
        raise HTTPException(status_code=404, detail="Annonce non trouvee")
    data = ad_update.dict(exclude_unset=True)
    for key, value in data.items():
        setattr(ad, key, value)
    ad.updated_at = datetime.now()
    session.add(ad)
    session.commit()
    session.refresh(ad)
    return ad


@router.delete("/ads/{ad_id}")
async def delete_ad(ad_id: int, session: Session = Depends(get_session)):
    """Supprime une annonce."""
    ad = session.get(Ad, ad_id)
    if not ad:
        raise HTTPException(status_code=404, detail="Annonce non trouvee")
    session.delete(ad)
    session.commit()
    return {"message": "Annonce supprimee"}


# ==================== IMPORT JSON ====================

@router.post("/contracts/import-json")
async def import_contracts_json(
    mode: str = Query("merge", description="Mode: 'merge' ou 'replace'"),
    file: Optional[UploadFile] = File(None),
    body: Optional[str] = None,
    session: Session = Depends(get_session)
):
    """
    Importe des contrats depuis un fichier JSON ou un body JSON.
    
    - file: Fichier JSON (multipart/form-data)
    - body: JSON direct (application/json)
    - mode: "merge" (met à jour) ou "replace" (remplace)
    """
    from app.services.contract_json_importer import import_contracts_from_json
    
    if mode not in ["merge", "replace"]:
        raise HTTPException(status_code=400, detail="mode doit être 'merge' ou 'replace'")
    
    json_data = None
    
    if file:
        # Lire depuis le fichier uploadé
        content = await file.read()
        try:
            json_data = json.loads(content.decode('utf-8'))
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"JSON invalide: {str(e)}")
    elif body:
        # Lire depuis le body
        try:
            json_data = json.loads(body)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"JSON invalide: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Fournir soit 'file' soit 'body' JSON")
    
    try:
        result = import_contracts_from_json(json_data, mode=mode, session=session)
        return {
            "message": "Import terminé",
            "imported": result["imported"],
            "updated": result["updated"],
            "errors": result["errors"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'import: {str(e)}")


# ==================== AUTH ====================

def decode_token(token: str) -> Optional[int]:
    """Décode un token simple et retourne l'user_id."""
    try:
        decoded = base64.b64decode(token).decode()
        user_id_str, timestamp = decoded.split(":")
        return int(user_id_str)
    except Exception:
        return None


def get_current_user(
    authorization: Optional[str] = Header(None),
    session: Session = Depends(get_session)
) -> Optional[User]:
    """Récupère l'utilisateur courant depuis le token."""
    if not authorization:
        return None
    
    # Format: "Bearer <token>"
    if not authorization.startswith("Bearer "):
        return None
    
    token = authorization[7:]
    user_id = decode_token(token)
    if not user_id:
        return None
    
    user = session.get(User, user_id)
    if not user or not user.is_active:
        return None
    
    return user


def require_admin(user: Optional[User] = Depends(get_current_user)) -> User:
    """Vérifie que l'utilisateur est un admin."""
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    if user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    return user


@router.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest, session: Session = Depends(get_session)):
    """Connexion utilisateur."""
    statement = select(User).where(User.username == request.username)
    user = session.exec(statement).first()
    
    if not user or not verify_password(request.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Identifiants incorrects")
    
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Compte désactivé")
    
    # Mettre à jour last_login
    user.last_login = datetime.now()
    session.add(user)
    session.commit()
    
    # Générer un token simple
    token_data = f"{user.id}:{datetime.now().timestamp()}"
    token = base64.b64encode(token_data.encode()).decode()
    
    return LoginResponse(
        user_id=user.id,
        username=user.username,
        display_name=user.display_name,
        role=user.role.value,
        linked_code_union=user.linked_code_union,
        linked_groupe=user.linked_groupe,
        avatar_url=user.avatar_url,
        token=token
    )


@router.get("/auth/me", response_model=UserResponse)
async def get_me(user: User = Depends(get_current_user)):
    """Récupère l'utilisateur courant."""
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    
    return UserResponse(
        id=user.id,
        username=user.username,
        display_name=user.display_name,
        role=user.role.value,
        linked_code_union=user.linked_code_union,
        linked_groupe=user.linked_groupe,
        avatar_url=user.avatar_url,
        is_active=user.is_active,
        created_at=user.created_at,
        last_login=user.last_login
    )


@router.post("/auth/logout")
async def logout():
    """Déconnexion (côté client, invalider le token)."""
    return {"message": "Déconnecté"}


# ==================== USERS (Admin only) ====================

@router.get("/users", response_model=List[UserResponse])
async def list_users(admin: User = Depends(require_admin), session: Session = Depends(get_session)):
    """Liste tous les utilisateurs (admin only)."""
    statement = select(User).order_by(User.role, User.username)
    users = session.exec(statement).all()
    return [
        UserResponse(
            id=u.id,
            username=u.username,
            display_name=u.display_name,
            role=u.role.value,
            linked_code_union=u.linked_code_union,
            linked_groupe=u.linked_groupe,
            avatar_url=u.avatar_url,
            is_active=u.is_active,
            created_at=u.created_at,
            last_login=u.last_login
        )
        for u in users
    ]


@router.post("/users", response_model=UserResponse)
async def create_user(
    user_data: UserCreate,
    admin: User = Depends(require_admin),
    session: Session = Depends(get_session)
):
    """Crée un nouvel utilisateur (admin only)."""
    # Vérifier que le username n'existe pas
    statement = select(User).where(User.username == user_data.username)
    existing = session.exec(statement).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ce nom d'utilisateur existe déjà")
    
    # Valider le rôle
    try:
        role = UserRole(user_data.role)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Rôle invalide: {user_data.role}")
    
    user = User(
        username=user_data.username,
        password_hash=hash_password(user_data.password),
        display_name=user_data.display_name,
        role=role,
        linked_code_union=user_data.linked_code_union.upper() if user_data.linked_code_union else None,
        linked_groupe=user_data.linked_groupe.upper() if user_data.linked_groupe else None,
        is_active=True
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    
    return UserResponse(
        id=user.id,
        username=user.username,
        display_name=user.display_name,
        role=user.role.value,
        linked_code_union=user.linked_code_union,
        linked_groupe=user.linked_groupe,
        avatar_url=user.avatar_url,
        is_active=user.is_active,
        created_at=user.created_at,
        last_login=user.last_login
    )


@router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_data: UserUpdate,
    admin: User = Depends(require_admin),
    session: Session = Depends(get_session)
):
    """Met à jour un utilisateur (admin only)."""
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    if user_data.display_name is not None:
        user.display_name = user_data.display_name
    if user_data.password:
        user.password_hash = hash_password(user_data.password)
    if user_data.role is not None:
        try:
            user.role = UserRole(user_data.role)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Rôle invalide: {user_data.role}")
    if user_data.linked_code_union is not None:
        user.linked_code_union = user_data.linked_code_union.upper() if user_data.linked_code_union else None
    if user_data.linked_groupe is not None:
        user.linked_groupe = user_data.linked_groupe.upper() if user_data.linked_groupe else None
    if user_data.avatar_url is not None:
        user.avatar_url = user_data.avatar_url
    if user_data.is_active is not None:
        user.is_active = user_data.is_active
    
    session.add(user)
    session.commit()
    session.refresh(user)
    
    return UserResponse(
        id=user.id,
        username=user.username,
        display_name=user.display_name,
        role=user.role.value,
        linked_code_union=user.linked_code_union,
        linked_groupe=user.linked_groupe,
        avatar_url=user.avatar_url,
        is_active=user.is_active,
        created_at=user.created_at,
        last_login=user.last_login
    )


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: int,
    admin: User = Depends(require_admin),
    session: Session = Depends(get_session)
):
    """Supprime un utilisateur (admin only)."""
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Ne pas supprimer son propre compte
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Impossible de supprimer son propre compte")
    
    session.delete(user)
    session.commit()
    return {"message": "Utilisateur supprimé"}


# ==================== IMAGE UPLOAD ====================

@router.post("/uploads/ads")
async def upload_ad_image(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin)
):
    """Upload une image pour une annonce (admin only)."""
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp", "image/svg+xml", "image/pjpeg"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Type non supporté: {file.content_type}")
    ext = os.path.splitext(file.filename)[1] if file.filename else ".png"
    filename = f"{uuid.uuid4().hex}{ext}"
    try:
        from app.services.supabase_storage import upload_image
        content = await file.read()
        url = upload_image(content, filename, "ads", file.content_type or "image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload: {str(e)}")
    return {"filename": filename, "url": url}


@router.get("/uploads/ads/{filename}")
async def get_ad_image(filename: str):
    """Récupère une image uploadée."""
    filepath = os.path.join(UPLOADS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Image non trouvée")
    
    return FileResponse(filepath)


# ==================== AVATAR UPLOAD ====================

@router.post("/uploads/avatars")
async def upload_avatar(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin)
):
    """Upload une photo de profil (admin only)."""
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp", "image/pjpeg"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Type non supporté: {file.content_type}")
    ext = os.path.splitext(file.filename)[1] if file.filename else ".png"
    filename = f"{uuid.uuid4().hex}{ext}"
    try:
        from app.services.supabase_storage import upload_image
        content = await file.read()
        url = upload_image(content, filename, "avatars", file.content_type or "image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload: {str(e)}")
    return {"filename": filename, "url": url}


@router.get("/uploads/avatars/{filename}")
async def get_avatar(filename: str):
    """Récupère une photo de profil."""
    filepath = os.path.join(AVATARS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Image non trouvée")
    
    return FileResponse(filepath)


# ==================== LOGO UPLOAD ====================

@router.post("/uploads/logos")
async def upload_logo(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin)
):
    """Upload le logo de l'entreprise (admin only)."""
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp", "image/svg+xml", "image/pjpeg"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Type non supporté: {file.content_type}")
    ext = os.path.splitext(file.filename)[1] if file.filename else ".png"
    filename = f"company_logo{ext}"
    try:
        from app.services.supabase_storage import upload_image
        content = await file.read()
        url = upload_image(content, filename, "logos", file.content_type or "image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload: {str(e)}")
    return {"filename": filename, "url": url}


@router.get("/uploads/logos/{filename}")
async def get_logo(filename: str):
    """Récupère le logo."""
    filepath = os.path.join(LOGOS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Image non trouvée")
    
    return FileResponse(filepath)


# ==================== SUPPLIER LOGOS ====================

@router.get("/supplier-logos")
async def get_supplier_logos(session: Session = Depends(get_session)):
    """Liste tous les logos fournisseurs."""
    logos = session.exec(select(SupplierLogo).order_by(SupplierLogo.supplier_key)).all()
    return [
        {
            "id": logo.id,
            "supplier_key": logo.supplier_key,
            "supplier_name": logo.supplier_name,
            "image_url": logo.image_url,
            "is_active": logo.is_active,
        }
        for logo in logos
    ]


@router.post("/supplier-logos")
async def create_supplier_logo(
    supplier_key: str = Form(...),
    supplier_name: str = Form(...),
    file: UploadFile = File(...),
    session: Session = Depends(get_session)
):
    """Upload un logo fournisseur."""
    # Verifier si le supplier_key existe deja
    existing = session.exec(
        select(SupplierLogo).where(SupplierLogo.supplier_key == supplier_key.upper())
    ).first()
    
    # Sauvegarder l'image (Supabase Storage sur Vercel, disque en local)
    ext = os.path.splitext(file.filename)[1] or ".png"
    filename = f"{supplier_key.upper()}{ext}"
    content = await file.read()
    try:
        from app.services.supabase_storage import upload_image
        image_url = upload_image(content, filename, "supplier_logos", file.content_type or "image/jpeg")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload: {str(e)}")
    
    if existing:
        existing.supplier_name = supplier_name
        existing.image_url = image_url
        existing.updated_at = datetime.now()
        session.add(existing)
        session.commit()
        session.refresh(existing)
        return {"id": existing.id, "supplier_key": existing.supplier_key, "supplier_name": existing.supplier_name, "image_url": existing.image_url, "is_active": existing.is_active}
    else:
        logo = SupplierLogo(
            supplier_key=supplier_key.upper(),
            supplier_name=supplier_name,
            image_url=image_url,
            is_active=True
        )
        session.add(logo)
        session.commit()
        session.refresh(logo)
        return {"id": logo.id, "supplier_key": logo.supplier_key, "supplier_name": logo.supplier_name, "image_url": logo.image_url, "is_active": logo.is_active}


@router.delete("/supplier-logos/{logo_id}")
async def delete_supplier_logo(logo_id: int, session: Session = Depends(get_session)):
    """Supprime un logo fournisseur."""
    logo = session.get(SupplierLogo, logo_id)
    if not logo:
        raise HTTPException(status_code=404, detail="Logo non trouve")
    
    # Supprimer le fichier
    if logo.image_url:
        filename = logo.image_url.split("/")[-1]
        filepath = os.path.join(SUPPLIER_LOGOS_DIR, filename)
        if os.path.exists(filepath):
            os.remove(filepath)
    
    session.delete(logo)
    session.commit()
    return {"ok": True}


@router.get("/uploads/supplier-logos/{filename}")
async def get_supplier_logo_file(filename: str):
    """Sert un fichier logo fournisseur."""
    filepath = os.path.join(SUPPLIER_LOGOS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Image non trouvee")
    return FileResponse(filepath)


# ==================== SETTINGS ====================

@router.get("/settings/{key}")
async def get_setting(key: str, session: Session = Depends(get_session)):
    """Récupère un paramètre."""
    statement = select(AppSettings).where(AppSettings.key == key)
    setting = session.exec(statement).first()
    if not setting:
        return {"key": key, "value": None}
    return {"key": setting.key, "value": setting.value}


@router.put("/settings/{key}")
async def set_setting(
    key: str,
    value: str = Query(...),
    admin: User = Depends(require_admin),
    session: Session = Depends(get_session)
):
    """Met à jour un paramètre (admin only)."""
    statement = select(AppSettings).where(AppSettings.key == key)
    setting = session.exec(statement).first()
    
    if setting:
        setting.value = value
        setting.updated_at = datetime.now()
    else:
        setting = AppSettings(key=key, value=value)
    
    session.add(setting)
    session.commit()
    session.refresh(setting)
    
    return {"key": setting.key, "value": setting.value}


@router.get("/settings")
async def list_settings(session: Session = Depends(get_session)):
    """Liste tous les paramètres."""
    statement = select(AppSettings)
    settings = session.exec(statement).all()
    return {s.key: s.value for s in settings}


# ==================== TEST IMPORT BRUT (ISOLÉ - NE MODIFIE PAS LE CODE EXISTANT) ====================

@router.post("/test/upload-raw")
async def test_upload_raw(file: UploadFile = File(...), year_filter: Optional[int] = None):
    """
    TEST UNIQUEMENT - Analyse un fichier brut sans rien sauvegarder.
    Ce endpoint est complètement isolé et ne modifie pas le code existant.
    Retourne un rapport détaillé de validation.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Le fichier doit être un .xlsx ou .xls")
    
    tmp_path = None
    try:
        # Sauvegarder temporairement
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        # Importer le service de test (isolé)
        from app.services.test_raw_import import (
            detect_file_format,
            load_excel_raw,
            validate_rules,
            calculate_rfa_from_raw,
            generate_validation_report
        )
        import pandas as pd
        
        # Détecter le format
        df = pd.read_excel(tmp_path, engine='openpyxl')
        raw_columns = list(df.columns)
        file_format = detect_file_format(raw_columns)
        
        if file_format == "large":
            # C'est un fichier FORMAT LARGE (déjà calculé par AppScript)
            # Utiliser le système existant pour le lire
            from app.services.excel_import import load_excel
            
            data, raw_columns_loaded, column_mapping = load_excel(tmp_path)
            
            # Générer un rapport pour format large
            report = {
                "file_format": "large",
                "message": "Fichier FORMAT LARGE détecté (déjà calculé par AppScript)",
                "summary": {
                    "total_lignes_brutes": len(data),
                    "colonnes_reconnues": len(column_mapping),
                    "colonnes_totales": len(raw_columns),
                },
                "columns_validation": {
                    "mapped_columns": column_mapping,
                    "all_raw_columns": raw_columns,
                    "columns_diagnostic": [
                        {
                            "field": key,
                            "label": label,
                            "found": key in column_mapping,
                            "excel_column": column_mapping.get(key),
                            "status": "✅ Reconnue" if key in column_mapping else "❌ Non reconnue"
                        }
                        for key, label, _ in [
                            ("code_union", "Code Union", []),
                            ("nom_client", "Nom Client", []),
                            ("groupe_client", "Groupe Client", []),
                            ("GLOBAL_ACR", "CA RFA GLOBALE ACR", []),
                            ("GLOBAL_ALLIANCE", "CA RFA GLOBALE ALLIANCE", []),
                            ("GLOBAL_DCA", "CA RFA GLOBALE DCA", []),
                            ("GLOBAL_EXADIS", "CA RFA GLOBALE EXADIS", []),
                        ]
                    ]
                },
                "note": "Ce fichier est au format LARGE (déjà calculé). Pour tester le calcul depuis le fichier BRUT, utilisez le fichier d'entrée du script AppScript (feuille 'global New')."
            }
        else:
            # C'est un fichier BRUT
            # Charger les données brutes (sans sauvegarder)
            data, raw_columns_loaded, column_mapping, detected_month, detected_year, mapping_method, load_stats = load_excel_raw(tmp_path)
            
            # Détecter l'année si non fournie
            if not year_filter:
                year_filter = detected_year or datetime.now().year
            
            # Validation des règles
            rules_validation = validate_rules()
            
            # Calcul des RFA (sans sauvegarder)
            rfa_results = calculate_rfa_from_raw(data, year_filter)
            
            # Générer un rapport détaillé
            report = generate_validation_report(
                data=data,
                rfa_results=rfa_results,
                rules_validation=rules_validation,
                column_mapping=column_mapping,
                raw_columns=raw_columns,
                year_filter=year_filter
            )
            
            # Ajouter les statistiques de chargement
            report["load_statistics"] = load_stats
            report["detected_year"] = detected_year
            report["years_in_file"] = list(load_stats.get("exemples_annees", []))
            
            # Ajouter la méthode de mapping utilisée
            report["mapping_method"] = mapping_method
            report["mapping_method_description"] = {
                "name": "Mapping par nom de colonne",
                "position": "Mapping par position (comme AppScript - colonnes 1,2,3,4,6,7,8,9,10,11)",
                "mixed": "Mapping mixte (nom + position pour les colonnes non reconnues)"
            }.get(mapping_method, "Inconnu")
        
        # Ajouter le format du fichier
        report["file_format"] = file_format
        
        # Ajouter les statistiques de chargement
        report["load_statistics"] = load_stats
        report["detected_year"] = detected_year
        report["years_in_file"] = list(load_stats.get("exemples_annees", []))
        
        return report

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Erreur dans test_upload_raw: {e}")
        print(error_trace)
        return {
            "error": str(e),
            "traceback": error_trace,
            "message": "Une erreur est survenue lors de l'analyse du fichier. Vérifiez les logs du serveur."
        }
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except:
                pass


@router.get("/pure-data/load-from-supabase")
async def load_pure_data_from_supabase_endpoint(
    year_current: Optional[int] = None,
    year_previous: Optional[int] = None,
    month: Optional[int] = None,
):
    """
    Charge les données Pure Data depuis Supabase (sans upload fichier).
    Utilisé quand Google Sheets a déjà été synchronisé vers Supabase.
    """
    try:
        import sys
        sys.setrecursionlimit(5000)  # hausse la limite pour les grandes listes JSON

        from app.services.pure_data_import import filter_rows, aggregate_rows, build_comparison
        from app.services.pure_data_supabase import read_pure_data_from_supabase, count_pure_data_rows

        if count_pure_data_rows() == 0:
            raise HTTPException(status_code=404, detail="Aucune donnée Pure Data dans Supabase. Cliquez sur 'Synchroniser depuis Sheets'.")

        # Lire directement depuis Supabase avec filtre pour ne charger que l'utile
        # Si filtre sur mois : on charge year_current + year_previous seulement
        # Sans filtre : on charge tout (nécessaire pour la vue globale)
        current_rows_db, _, _ = read_pure_data_from_supabase(year=year_current, month=month)
        previous_rows_db, _, _ = read_pure_data_from_supabase(year=year_previous, month=month)

        current_agg  = aggregate_rows(current_rows_db)
        previous_agg = aggregate_rows(previous_rows_db)
        comparison   = build_comparison(current_agg, previous_agg)

        # Limiter les listes clients/commerciaux pour éviter des réponses JSON trop lourdes
        MAX_ITEMS = 500
        if "clients" in comparison:
            comparison["clients"] = comparison["clients"][:MAX_ITEMS]
        if "commercials" in comparison:
            comparison["commercials"] = comparison["commercials"][:MAX_ITEMS]

        return {
            "pure_data_id": "sheets_live",
            "current":  {"year": year_current, "month": month, "total_ca": current_agg["total_ca"], "row_count": len(current_rows_db)},
            "previous": {"year": year_previous, "month": month, "total_ca": previous_agg["total_ca"], "row_count": len(previous_rows_db)},
            "comparison": comparison,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback; print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur chargement Pure Data: {str(e)}")


@router.get("/pure-data/sheets-status")
async def pure_data_sheets_status():
    """Vérifie si des données Pure Data sont disponibles dans Supabase."""
    try:
        from app.services.pure_data_supabase import count_pure_data_rows
        count = count_pure_data_rows()
        return {
            "has_data": count > 0,
            "row_count": count,
            "sheet_name": __import__("os").environ.get("PURE_DATA_SHEET_NAME", "global New"),
            "spreadsheet_id": __import__("os").environ.get("RFA_SHEETS_SPREADSHEET_ID", "16Hog9Dc43vwj_JmjRBLlIPaYoHoxLKVB7eSrBVXOLM0"),
        }
    except Exception as e:
        return {"has_data": False, "row_count": 0, "error": str(e)}


@router.post("/pure-data/sync-sheets")
async def sync_pure_data_from_sheets(admin: User = Depends(require_admin)):
    """
    Charge la feuille 'global New' depuis Google Sheets et stocke dans Supabase.
    Déclenché manuellement par un admin.
    """
    try:
        from app.services.pure_data_sheets import load_pure_data_from_sheets
        from app.services.pure_data_supabase import write_pure_data_to_supabase

        rows, columns, mapping = load_pure_data_from_sheets()
        if not rows:
            raise HTTPException(status_code=400, detail="Aucune donnée trouvée dans la feuille Google Sheets")

        n = write_pure_data_to_supabase(rows)

        # Invalider le cache mémoire
        from app.storage import _pure_data_imports
        _pure_data_imports.pop("sheets_live", None)

        return {
            "success": True,
            "rows_imported": n,
            "columns": columns[:10],
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur sync Pure Data Sheets: {str(e)}")


def _resolve_pure_data(pure_data_id: str):
    """
    Résout un import Pure Data : depuis la mémoire d'abord, puis Supabase si sheets_live.
    """
    from app.storage import get_pure_data_import, create_pure_data_import, _pure_data_imports
    # 1. Mémoire (cache)
    pd_import = _resolve_pure_data(pure_data_id)
    if pd_import:
        return pd_import
    # 2. Supabase (pour sheets_live ou après redémarrage)
    if pure_data_id == "sheets_live":
        try:
            from app.services.pure_data_supabase import read_pure_data_from_supabase
            # Charger TOUTES les données (le filtrage se fait ensuite en Python)
            # sys.setrecursionlimit est géré dans l'endpoint appelant
            rows, columns, mapping = read_pure_data_from_supabase()
            if rows:
                _id = create_pure_data_import(columns, mapping, rows)
                _pure_data_imports["sheets_live"] = _pure_data_imports[_id]
                return _pure_data_imports["sheets_live"]
        except Exception as e:
            print(f"[PURE DATA] Erreur lecture Supabase: {e}")
    return None


@router.post("/pure-data/compare")
async def compare_pure_data(
    file: UploadFile = File(...),
    year_current: Optional[int] = Form(None),
    year_previous: Optional[int] = Form(None),
    month: Optional[int] = Form(None)
):
    """
    Compare des données "pure data" (N vs N-1) à partir de 2 fichiers Excel.
    Filtre optionnel sur l'année et le mois.
    """
    from app.services.pure_data_import import load_pure_data, filter_rows, aggregate_rows, build_comparison

    tmp_current = None
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Fichier .xlsx ou .xls requis")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(await file.read())
            tmp_current = tmp.name
        current_rows, current_cols, current_mapping = load_pure_data(tmp_current)
        pure_data_id = create_pure_data_import(current_cols, current_mapping, current_rows)

        current_filtered  = filter_rows(current_rows, year_current, month)
        previous_filtered = filter_rows(current_rows, year_previous, month)
        current_agg  = aggregate_rows(current_filtered)
        previous_agg = aggregate_rows(previous_filtered)
        comparison   = build_comparison(current_agg, previous_agg)

        return {
            "pure_data_id": pure_data_id,
            "current": {
                "year": year_current, "month": month,
                "total_ca": current_agg["total_ca"],
                "row_count": len(current_filtered),
                "columns": current_cols, "mapping": current_mapping,
            },
            "previous": {
                "year": year_previous, "month": month,
                "total_ca": previous_agg["total_ca"],
                "row_count": len(previous_filtered),
                "columns": current_cols, "mapping": current_mapping,
            },
            "comparison": comparison,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur analyse pure data: {str(e)}")
    finally:
        if tmp_current and os.path.exists(tmp_current):
            try:
                os.unlink(tmp_current)
            except Exception:
                pass


@router.get("/pure-data/comparison")
async def get_pure_data_comparison(
    pure_data_id: str,
    year_current: Optional[int] = None,
    year_previous: Optional[int] = None,
    month: Optional[int] = None,
    fournisseur: Optional[str] = None,
):
    """
    Recalcule la comparaison N vs N-1 à partir des lignes stockées, avec filtre fournisseur optionnel.
    Utilisé pour afficher Pure Data filtré par plateforme (ACR, DCA, etc.).
    """
    pure_data = _resolve_pure_data(pure_data_id)
    if not pure_data:
        raise HTTPException(status_code=404, detail="Import pure data introuvable. Relance l'analyse.")
    from app.services.pure_data_import import (
        filter_rows,
        filter_rows_by_fournisseur,
        aggregate_rows,
        build_comparison,
    )
    rows = pure_data.rows
    current_filtered = filter_rows(rows, year_current, month)
    current_filtered = filter_rows_by_fournisseur(current_filtered, fournisseur)
    previous_filtered = filter_rows(rows, year_previous, month)
    previous_filtered = filter_rows_by_fournisseur(previous_filtered, fournisseur)
    current_agg = aggregate_rows(current_filtered)
    previous_agg = aggregate_rows(previous_filtered)
    comparison = build_comparison(current_agg, previous_agg)
    return {"comparison": comparison}


@router.get("/pure-data/client-detail")
async def pure_data_client_detail(
    pure_data_id: str,
    code_union: str,
    year_current: Optional[int] = None,
    year_previous: Optional[int] = None,
    month: Optional[int] = None,
    fournisseur: Optional[str] = None,
):
    """
    Détail N vs N-1 pour un client (fournisseur -> marque -> famille -> sous-famille).
    Si fournisseur est fourni, seul ce fournisseur est inclus.
    """
    pure_data = _resolve_pure_data(pure_data_id)
    if not pure_data:
        raise HTTPException(status_code=404, detail="Import pure data introuvable. Relance l'analyse.")

    from app.services.pure_data_import import build_client_detail
    detail = build_client_detail(
        pure_data.rows,
        code_union=code_union,
        year_current=year_current,
        year_previous=year_previous,
        month=month,
        fournisseur=fournisseur,
    )
    return detail


@router.get("/pure-data/platform-detail")
async def pure_data_platform_detail(
    pure_data_id: str,
    platform: str,
    year_current: Optional[int] = None,
    year_previous: Optional[int] = None,
    month: Optional[int] = None
):
    """
    Détail N vs N-1 pour une plateforme (liste clients).
    """
    pure_data = _resolve_pure_data(pure_data_id)
    if not pure_data:
        raise HTTPException(status_code=404, detail="Import pure data introuvable. Relance l'analyse.")

    from app.services.pure_data_import import build_platform_detail
    try:
        detail = build_platform_detail(
            pure_data.rows,
            fournisseur=platform,
            year_current=year_current,
            year_previous=year_previous,
            month=month
        )
        return detail
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur détail plateforme: {str(e)}")


@router.get("/pure-data/marque-detail")
async def pure_data_marque_detail(
    pure_data_id: str,
    platform: str,
    marque: str,
    year_current: Optional[int] = None,
    year_previous: Optional[int] = None,
    month: Optional[int] = None,
):
    """
    Pour une plateforme et une marque, retourne les magasins (clients) qui contribuent à cette marque.
    """
    pure_data = _resolve_pure_data(pure_data_id)
    if not pure_data:
        raise HTTPException(status_code=404, detail="Import pure data introuvable. Relance l'analyse.")

    from app.services.pure_data_import import build_marque_detail
    try:
        detail = build_marque_detail(
            pure_data.rows,
            fournisseur=platform,
            marque=marque,
            year_current=year_current,
            year_previous=year_previous,
            month=month,
        )
        return detail
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur détail marque: {str(e)}")


@router.get("/pure-data/commercial-detail")
async def pure_data_commercial_detail(
    pure_data_id: str,
    commercial: str,
    year_current: Optional[int] = None,
    year_previous: Optional[int] = None,
    month: Optional[int] = None,
    fournisseur: Optional[str] = None,
):
    """
    Détail N vs N-1 pour un commercial (global + plateformes + clients).
    """
    pure_data = _resolve_pure_data(pure_data_id)
    if not pure_data:
        raise HTTPException(status_code=404, detail="Import pure data introuvable. Relance l'analyse.")

    from app.services.pure_data_import import build_commercial_detail
    detail = build_commercial_detail(
        pure_data.rows,
        commercial=commercial,
        year_current=year_current,
        year_previous=year_previous,
        month=month,
        fournisseur=fournisseur,
    )
    return detail


# ==================== GENIE RFA (Assistant commercial IA) ====================

@router.get("/genie/query")
async def genie_query_endpoint(
    import_id: str,
    query_type: str,
    key: Optional[str] = None,
    search: Optional[str] = None,
    limit: Optional[int] = 10,
    session: Session = Depends(get_session),
):
    """
    Chatbot Génie RFA : requêtes prédéfinies.
    query_type: dashboard, top_gains, near_by_objective, union_opportunities, search_adherent, entity_profile, smart_plan, cascade, double_lever, balance
    search: pour search_adherent, entity_profile, smart_plan (nom ou code Union).
    """
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        raise HTTPException(status_code=404, detail="Import non trouvé")

    from app.services.genie_engine import genie_query, genie_query_fast
    params = {}
    if key:
        params["key"] = key
    if search:
        params["search"] = search
    if limit:
        params["limit"] = limit

    is_vercel = os.environ.get("VERCEL") == "1"

    try:
        if is_vercel:
            # Vercel free tier : timeout 10s, version rapide (CA, sans résolution contrat)
            result = genie_query_fast(import_data, query_type, params)
            result["_limited_mode"] = True
            result["_note"] = "Analyse CA (cloud). Ouvrez l'application locale pour l'analyse complète avec objectifs et contrats."
            return result
        else:
            return genie_query(import_data, query_type, params)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur Génie: {str(e)}")


@router.get("/genie/smart-plans")
async def genie_smart_plans(import_id: str, entity_id: Optional[str] = None, session: Session = Depends(get_session)):
    """
    Retourne les plans d'achat optimisés.
    Si entity_id fourni : plans de ce client uniquement.
    Sinon : tous les plans.
    """
    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        raise HTTPException(status_code=404, detail="Import non trouvé")
    from app.services.genie_engine import genie_full_analysis
    try:
        analysis = genie_full_analysis(import_data)
        plans = analysis.get("smart_plans", [])
        if entity_id:
            search = entity_id.upper()
            plans = [p for p in plans if search in p["entity_id"].upper() or search in p["entity_label"].upper()]
        return plans
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur plans: {str(e)}")


@router.get("/genie/smart-plans/export-excel")
async def genie_export_excel(import_id: str, session: Session = Depends(get_session)):
    """Export Excel structuré de tous les plans d'achat optimisés."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    import_data = _resolve_import_data(import_id, session)
    if not import_data:
        raise HTTPException(status_code=404, detail="Import non trouvé")
    from app.services.genie_engine import genie_full_analysis
    try:
        analysis = genie_full_analysis(import_data)
        plans = analysis.get("smart_plans", [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur analyse: {str(e)}")

    wb = openpyxl.Workbook()

    # --- Feuille 1 : Synthèse Plans ---
    ws = wb.active
    ws.title = "Plans d'achat"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    bonus_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    green_font = Font(bold=True, color="047857")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Code Union", "Nom Client", "Fournisseur", "CA Global actuel",
        "Palier Global visé", "Manque Global",
        "Paliers débloqués", "Paliers avec bonus", "CA à investir",
        "CA avec bonus", "Gain RFA Option A", "Gain RFA Option B",
        "Bonus effort", "Tri-partites à pousser"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin

    row_num = 2
    for plan in plans:
        tri_detail = " | ".join(
            f"{it['label']}: +{it['ca_to_push']:,.0f}€ → +{it['projected_gain']:,.0f}€ RFA"
            for it in plan["plan_items"]
        )
        vals = [
            plan["entity_id"],
            plan["entity_label"].split(" - ", 1)[1] if " - " in plan["entity_label"] else plan["entity_label"],
            plan["global_label"],
            plan["global_ca"],
            plan["global_ca"] + plan["global_missing"],
            plan["global_missing"],
            plan["tiers_unlocked"],
            plan.get("tiers_with_bonus", plan["tiers_unlocked"]),
            plan["total_ca_needed"],
            plan.get("total_with_bonus", plan["total_ca_needed"]),
            plan.get("gain_option_a", 0),
            plan.get("gain_option_b", 0),
            plan.get("bonus_effort", 0),
            tri_detail,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin
            if col in (4, 5, 6, 9, 10, 11, 12, 13):
                cell.number_format = '#,##0'
            if col == 11:
                cell.font = green_font
            if col == 12:
                cell.font = Font(bold=True, color="0E7490")
        row_num += 1

    # Ajuster largeur
    col_widths = [12, 30, 20, 15, 15, 12, 10, 10, 14, 14, 14, 14, 12, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Feuille 2 : Détail tri-partites ---
    ws2 = wb.create_sheet("Détail tri-partites")
    headers2 = [
        "Code Union", "Nom Client", "Fournisseur",
        "Tri-partite", "CA actuel tri", "Progression %",
        "CA à pousser", "Gain RFA tri",
        "Contribue au global", "Global débloqué?"
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin

    row_num = 2
    for plan in plans:
        nom = plan["entity_label"].split(" - ", 1)[1] if " - " in plan["entity_label"] else plan["entity_label"]
        for item in plan["plan_items"]:
            vals2 = [
                plan["entity_id"],
                nom,
                plan["global_label"],
                item["label"],
                item["ca"],
                round(item["progress"], 1),
                item["ca_to_push"],
                item["projected_gain"],
                "Oui",
                "Oui" if plan["global_unlocked"] else ("Avec bonus" if plan.get("bonus_reasonable") else "Non"),
            ]
            for col, v in enumerate(vals2, 1):
                cell = ws2.cell(row=row_num, column=col, value=v)
                cell.border = thin
                if col in (5, 7, 8):
                    cell.number_format = '#,##0'
            row_num += 1

    col_widths2 = [12, 30, 20, 25, 14, 10, 14, 14, 15, 15]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Plans_Achat_RFA.xlsx"'}
    )


# ═══════════════════════════════════════════════════════════════════
#  NATHALIE — Ouverture de comptes
# ═══════════════════════════════════════════════════════════════════

@router.get("/nathalie/clients")
async def nathalie_clients(ouverture_only: bool = False):
    """
    Liste des clients depuis LISTE CLIENT 2.
    ouverture_only=true : uniquement ceux avec OUVERTURE CHEZ renseigné.
    """
    try:
        clients = nathalie_service.get_clients(with_ouverture_only=ouverture_only)
        return {"clients": clients, "total": len(clients)}
    except ValueError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture Sheet : {str(e)}")


@router.get("/nathalie/suppliers")
async def nathalie_suppliers():
    """Liste des contacts fournisseurs depuis CONTACT FOURNISSEURS."""
    try:
        suppliers = nathalie_service.get_suppliers()
        return {"suppliers": suppliers, "total": len(suppliers)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture Sheet : {str(e)}")


@router.get("/nathalie/tasks")
async def nathalie_tasks(code_union: Optional[str] = None):
    """Tâches depuis TACHE CLIENTS, optionnellement filtrées par code union."""
    try:
        tasks = nathalie_service.get_tasks(code_union=code_union)
        return {"tasks": tasks, "total": len(tasks)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lecture Sheet : {str(e)}")


@router.post("/nathalie/generate-emails")
async def nathalie_generate_emails(body: Dict[str, Any] = Body(...)):
    """
    Génère un email par fournisseur (confidentialité : chaque fournisseur ne voit que son propre mail).
    Body: { "code_union": "M0024", "supplier_names": ["ACR", "DCA"] }
    Destinataire = contact du fournisseur (feuille CONTACT FOURNISSEURS).
    """
    code_union: str = body.get("code_union", "")
    supplier_names: List[str] = body.get("supplier_names", [])

    if not code_union:
        raise HTTPException(status_code=400, detail="code_union requis")

    try:
        client = nathalie_service.get_client_by_code(code_union)
        if not client:
            raise HTTPException(status_code=404, detail=f"Client {code_union} introuvable")

        all_suppliers = nathalie_service.get_suppliers()
        sup_map: Dict[str, Any] = {
            s["entreprise"].upper(): s for s in all_suppliers if s.get("entreprise")
        }

        emails = []
        for name in supplier_names:
            supplier = sup_map.get(name.upper())
            if not supplier:
                # Fournisseur sans contact connu — email générique vide
                supplier = {"entreprise": name, "nom": "", "prenom": "", "mail": "", "telephone": "", "poste": ""}
            email = nathalie_service.generate_email(client, supplier)
            emails.append(email)

        return {"emails": emails, "client": client}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur génération email : {str(e)}")


@router.post("/nathalie/send-emails")
async def nathalie_send_emails(body: Dict[str, Any] = Body(...)):
    """
    Envoie réellement les emails aux fournisseurs via l'API Gmail (un mail par fournisseur).
    Pièces jointes : RIB, Kbis, pièce d'identité (téléchargées depuis Drive).
    Body: { "code_union": "M0024", "supplier_names": ["ACR", "DCA"], "cc_emails": ["a@groupementunion.pro"] }
    cc_emails optionnel ; sinon pris depuis la variable d'environnement NATHALIE_CC_EMAILS.
    """
    code_union: str = body.get("code_union", "")
    supplier_names: List[str] = body.get("supplier_names", [])
    cc_emails: Optional[List[str]] = body.get("cc_emails")

    if not code_union:
        raise HTTPException(status_code=400, detail="code_union requis")
    if not supplier_names:
        raise HTTPException(status_code=400, detail="supplier_names requis")

    try:
        results = nathalie_service.send_emails_to_suppliers(
            code_union=code_union,
            supplier_names=supplier_names,
            cc_emails=cc_emails,
        )
        sent = sum(1 for r in results if r.get("success"))
        return {
            "results": results,
            "sent": sent,
            "total": len(results),
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur envoi email : {str(e)}")


@router.get("/nathalie/client/{code_union}")
async def nathalie_client_detail(code_union: str):
    """Détail complet d'un client + tâches associées."""
    try:
        client = nathalie_service.get_client_by_code(code_union)
        if not client:
            raise HTTPException(status_code=404, detail=f"Client {code_union} introuvable")
        tasks = nathalie_service.get_tasks(code_union=code_union)
        return {"client": client, "tasks": tasks}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/nathalie/create-client")
async def nathalie_create_client(
    # Champs formulaire
    nom_client: str = Form(...),
    groupe: str = Form(...),
    adresse: str = Form(""),
    code_postal: str = Form(""),
    ville: str = Form(""),
    telephone: str = Form(""),
    mail: str = Form(""),
    siret: str = Form(""),
    agent_union: str = Form(""),
    contrat_type: str = Form(""),
    notes: str = Form(""),
    # Fichiers
    rib: UploadFile = File(None),
    kbis: UploadFile = File(None),
    piece_identite: UploadFile = File(None),
):
    """
    Crée un client complet :
    1. Génère Code Union
    2. Crée dossier Drive
    3. Upload fichiers
    4. Ajoute ligne Sheet
    """
    data = {
        "nom_client": nom_client,
        "groupe": groupe,
        "adresse": adresse,
        "code_postal": code_postal,
        "ville": ville,
        "telephone": telephone,
        "mail": mail,
        "siret": siret,
        "agent_union": agent_union,
        "contrat_type": contrat_type,
        "notes": notes,
    }
    files = {
        "rib": rib,
        "kbis": kbis,
        "piece_identite": piece_identite,
    }

    try:
        result = await nathalie_service.create_client_full(data, files)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur création client : {str(e)}")


